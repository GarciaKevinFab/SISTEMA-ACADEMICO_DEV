# backend/catalogs/views.py
from __future__ import annotations

import os
import io
import csv
import json
import re
import time
import zipfile
import unicodedata
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple
from datetime import timedelta
from django.utils import timezone
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.management import call_command
from django.db import models, transaction, IntegrityError
from django.db.utils import OperationalError
from django.http import Http404, HttpResponse, FileResponse
from django.db.models import Q
from django.contrib.auth import get_user_model
from django.utils.crypto import get_random_string
from openpyxl import Workbook, load_workbook
from rest_framework import viewsets
from rest_framework.decorators import action, api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from catalogs.models import Period
from acl.models import Role, UserRole
from students.models import Student
from academic.models import Career, Plan, Course, PlanCourse

# ✅ si implementaste AcademicGradeRecord
try:
    from academic.models import AcademicGradeRecord
except Exception:
    AcademicGradeRecord = None

# ✅ sincronizar carreras al módulo de admisión
try:
    from admission.models import Career as AdmissionCareer
except Exception:
    AdmissionCareer = None

from .models import (
    Campus, Classroom, Teacher,
    InstitutionSetting, MediaAsset,
    ImportJob, BackupExport
)
from .serializers import (
    PeriodSerializer, CampusSerializer, ClassroomSerializer, TeacherSerializer,
    MediaAssetSerializer, BackupExportSerializer
)

User = get_user_model()

# ─────────────────────────────────────────────────────────────
# Utils
# ─────────────────────────────────────────────────────────────

def _norm(s):
    s = "" if s is None else str(s)
    s = s.strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = re.sub(r"\s+", " ", s)
    return s

def _to_int(v):
    try:
        if v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None
        return int(float(s))
    except Exception:
        return None

def _to_float(v):
    """Convierte a float seguro, acepta coma como decimal"""
    try:
        if v is None:
            return None
        s = str(v).strip().replace(",", ".")
        if s == "":
            return None
        return float(s)
    except Exception:
        return None

def _parse_date_yyyy_mm_dd(s):
    s = (s or "").strip()
    if not s:
        return None
    try:
        y, m, d = [int(x) for x in s.split("-")]
        return date(y, m, d)
    except Exception:
        return None

def _retry_db(fn, tries=10, delay=0.5):
    """
    Reintenta operaciones DB cuando SQLite se bloquea.
    """
    last = None
    for _ in range(tries):
        try:
            return fn()
        except OperationalError as e:
            last = e
            if "database is locked" in str(e).lower():
                time.sleep(delay)
                continue
            raise
    if last:
        raise last

def list_items(serializer_cls, queryset):
    return Response({"items": serializer_cls(queryset, many=True).data})

def _require_staff(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return Response({"detail": "No autorizado."}, status=403)
    return None

def _require_staff_or_teacher(request):
    if not request.user.is_authenticated:
        return Response({"detail": "No autorizado."}, status=403)

    # staff OK
    if getattr(request.user, "is_staff", False):
        return None

    # teacher role OK (tu User tiene ManyToMany roles)
    try:
        if request.user.roles.filter(name__iexact="TEACHER").exists():
            return None
    except Exception:
        pass

    return Response({"detail": "No autorizado."}, status=403)

def _xlsx_response(wb: Workbook, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

def _csv_bytes(rows: List[dict], headers: List[str]) -> bytes:
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=headers, extrasaction="ignore")
    w.writeheader()
    for r in rows:
        rr = {}
        for k in headers:
            v = r.get(k)
            if isinstance(v, (dict, list)):
                rr[k] = json.dumps(v, ensure_ascii=False)
            else:
                rr[k] = "" if v is None else v
        w.writerow(rr)
    return out.getvalue().encode("utf-8-sig")

def _parse_period_code(code: str):
    """
    Acepta: 2026-I, 2026-II, 2026-III
    Retorna: (year:int, term:str) o (None, None)
    """
    s = (code or "").strip().upper()
    if not s:
        return None, None
    parts = s.split("-")
    if len(parts) != 2:
        return None, None
    try:
        year = int(parts[0])
    except Exception:
        return None, None
    term = parts[1]
    if term not in ("I", "II", "III"):
        return None, None
    return year, term

def _slug_code(s: str, maxlen: int = 10) -> str:
    s = _norm(s).upper()
    s = re.sub(r"[^A-Z0-9]", "", s)
    return (s[:maxlen] or "CRS")

# ─────────────────────────────────────────────────────────────
# Student headers (excel oficial)
# ─────────────────────────────────────────────────────────────

STUDENT_HEADER_ALIASES = {
    "region": "region",
    "provincia": "provincia",
    "distrito": "distrito",
    "codigo modular": "codigo_modular",
    "codigo_modular": "codigo_modular",
    "nombre de la institucion": "nombre_institucion",
    "nombre institucion": "nombre_institucion",
    "gestion": "gestion",
    "tipo": "tipo",
    "programa / carrera": "programa_carrera",
    "programa carrera": "programa_carrera",
    "ciclo": "ciclo",
    "turno": "turno",
    "seccion": "seccion",
    "periodo": "periodo",
    "apellido paterno": "apellido_paterno",
    "apellido materno": "apellido_materno",
    "nombres": "nombres",
    "fecha nac": "fecha_nac",
    "fecha nac.": "fecha_nac",
    "sexo": "sexo",
    "num documento": "num_documento",
    "num. documento": "num_documento",
    "lengua": "lengua",
    "discapacidad": "discapacidad",
    "tipo de discapacidad": "tipo_discapacidad",
}

def _read_rows(file, mapping: dict):
    """
    Lee XLSX/CSV. Devuelve lista de dicts con __row__.
    Auto-mapea el Excel oficial de alumnos si detecta encabezados.
    """
    name = (getattr(file, "name", "") or "").lower()

    def apply_mapping(row: dict):
        if not mapping:
            return row
        out = dict(row)
        for field, col in (mapping or {}).items():
            if col and col in row:
                out[field] = row.get(col)
        return out

    # XLSX
    if name.endswith((".xlsx", ".xlsm", ".xltx", ".xls")):
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]

        header_map = {}
        for j, h in enumerate(headers):
            nh = _norm(h)
            if nh in STUDENT_HEADER_ALIASES:
                header_map[j] = STUDENT_HEADER_ALIASES[nh]

        out = []
        for i, r in enumerate(rows[1:], start=2):
            row_raw = {headers[j]: r[j] for j in range(len(headers)) if headers[j]}
            row_raw = {k: ("" if v is None else v) for k, v in row_raw.items()}

            row_auto = {}
            for j in range(len(headers)):
                key = header_map.get(j)
                if key:
                    row_auto[key] = "" if r[j] is None else r[j]

            row_final = row_auto if row_auto else row_raw
            row_final = apply_mapping(row_final)
            out.append({**row_final, "__row__": i})
        return out

    # CSV
    raw = file.read()
    try:
        text = raw.decode("utf-8-sig")
    except Exception:
        text = raw.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    out = []
    for idx, row in enumerate(reader, start=2):
        row = {k.strip(): ("" if v is None else str(v).strip()) for k, v in row.items() if k}

        row_auto = {}
        for k, v in row.items():
            nk = _norm(k)
            if nk in STUDENT_HEADER_ALIASES:
                row_auto[STUDENT_HEADER_ALIASES[nk]] = v

        row_final = row_auto if row_auto else row
        row_final = apply_mapping(row_final)
        out.append({**row_final, "__row__": idx})
    return out

# ─────────────────────────────────────────────────────────────
# PLAN de estudios (tu Excel real: semestre como BLOQUE)
# ─────────────────────────────────────────────────────────────

def _career_from_sheet_name(sheet: str) -> str:
    raw = (sheet or "").strip()
    raw = re.sub(r"\b20\d{2}\b", "", raw).strip()
    raw = re.sub(r"\s+", " ", raw).strip()
    key = _norm(raw)

    MAP = {
        "inicial": "EDUCACIÓN INICIAL",
        "primaria": "EDUCACIÓN PRIMARIA",
        "ed. fisica": "EDUCACIÓN FÍSICA",
        "ed fisica": "EDUCACIÓN FÍSICA",
        "educacion fisica": "EDUCACIÓN FÍSICA",
        "comunicacion": "COMUNICACIÓN",
    }
    return MAP.get(key, raw.upper())

def _normalize_semester_value(v) -> str:
    s = "" if v is None else str(v)
    s = re.sub(r"\s+", "", s, flags=re.UNICODE)  # quita espacios, saltos, tabs, etc.
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.lower()

def _is_semester_label(v) -> Optional[int]:
    key = _normalize_semester_value(v)
    if not key:
        return None

    sem_map = {
        "primero": 1, "segundo": 2, "tercero": 3, "cuarto": 4, "quinto": 5,
        "sexto": 6, "septimo": 7, "octavo": 8, "noveno": 9, "decimo": 10,
        "1": 1, "2": 2, "3": 3, "4": 4, "5": 5, "6": 6, "7": 7, "8": 8, "9": 9, "10": 10,
        "i": 1, "ii": 2, "iii": 3, "iv": 4, "v": 5, "vi": 6, "vii": 7, "viii": 8, "ix": 9, "x": 10,
    }

    if key in sem_map:
        return sem_map[key]

    if key.startswith("semestre"):
        tail = key.replace("semestre", "")
        return sem_map.get(tail)

    return None

def _read_study_plan_xlsx(file):
    wb = load_workbook(file, data_only=True, read_only=True, keep_links=False)
    out = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        career_name = _career_from_sheet_name(sheet)
        if not career_name:
            continue

        current_sem = None
        course_col = None
        cred_col = None
        hours_col = None
        header_found = False

        # en vez de cortar, solo usamos esto para “salir” si realmente ya no hay nada
        empty_streak = 0

        for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_norm = [(_norm(x) if x is not None else "") for x in row]

            # ------------------------------------------------
            # 1) Header: puede repetirse en medio del documento
            # ------------------------------------------------
            is_header_row = any(("areas" in h) or ("asignaturas" in h) for h in row_norm) and any(("cred" in h) for h in row_norm)
            if is_header_row:
                # Re-detectar columnas cada vez que aparezca header
                course_col = None
                cred_col = None
                hours_col = None

                for j, h in enumerate(row_norm):
                    if ("areas" in h) or ("asignaturas" in h):
                        course_col = j
                        break
                for j, h in enumerate(row_norm):
                    if ("cred" in h) or ("credito" in h) or ("creditos" in h):
                        cred_col = j
                        break
                for j, h in enumerate(row_norm):
                    if ("hora" in h) or ("horas" in h) or (h == "nota"):
                        hours_col = j
                        break

                header_found = course_col is not None and cred_col is not None
                empty_streak = 0
                continue

            # Si aún no encontramos header, seguimos
            if not header_found:
                continue

            # ------------------------------------------------
            # 2) Detectar semestre (P R I M E R O, NOVENO, etc.)
            # ------------------------------------------------
            found_sem = None
            for cell in row:
                sem = _is_semester_label(cell)
                if sem:
                    found_sem = sem
                    break

            if found_sem:
                current_sem = int(found_sem)
                empty_streak = 0
                # OJO: esta fila puede traer también curso, así que NO continue

            if not current_sem:
                continue

            # ------------------------------------------------
            # 3) Leer curso
            # ------------------------------------------------
            if course_col is None or course_col >= len(row):
                continue

            course_name = "" if row[course_col] is None else str(row[course_col]).strip()

            # filas completamente vacías -> solo cuenta y sigue (NO break)
            if not course_name:
                # si la fila está realmente vacía completa, sube contador
                if all((c is None or str(c).strip() == "") for c in row):
                    empty_streak += 1
                else:
                    empty_streak = 0
                # si hay demasiados vacíos y ya estamos muy abajo, recién ahí cortamos
                if empty_streak >= 250:
                    break
                continue

            empty_streak = 0

            # Filas que NO son cursos
            bad = _norm(course_name)
            if bad in ("sem", "n", "n°", "no", "numero", "areas/asignaturas", "areas / asignaturas", "asignaturas", "total", "totales"):
                continue

            credits = 0
            if cred_col is not None and cred_col < len(row):
                credits = _to_int(row[cred_col], 0) or 0

            hours = 0
            if hours_col is not None and hours_col < len(row):
                hours = _to_int(row[hours_col], 0) or 0

            out.append({
                "__row__": ridx,
                "career_name": career_name,
                "semester": int(current_sem),
                "course_name": course_name,
                "credits": int(credits),
                "hours": int(hours),
            })

    return out

# ─────────────────────────────────────────────────────────────
# CALIFICACIONES.xlsx (real)
# ─────────────────────────────────────────────────────────────

def _read_calificaciones_xlsx(file):
    wb = load_workbook(file, data_only=True, read_only=True, keep_links=False)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [(_norm(h) if h is not None else "") for h in rows[0]]

    def idx_match(*names):
        wanted = {_norm(x) for x in names}
        for i, h in enumerate(headers):
            if h in wanted:
                return i
        return None

    i_doc = idx_match("NUMERO_DOCUMENTO", "NUM DOCUMENTO", "NUM_DOCUMENTO", "DNI")
    i_periodo = idx_match("PERIODO", "PERIOD")
    i_programa = idx_match("PROGRAMA", "PROGRAMA / CARRERA", "PROGRAMA_CARRERA")
    i_ciclo = idx_match("CICLO")
    i_curso = idx_match("CURSO", "ASIGNATURA")
    i_nota = idx_match("PROMEDIO_VIGESIMAL", "PROMEDIO", "FINAL", "NOTA", "PROM_FINAL")

    out = []
    for ridx, r in enumerate(rows[1:], start=2):
        def get(i):
            return r[i] if i is not None and i < len(r) else None

        doc = "" if get(i_doc) is None else str(get(i_doc)).strip()
        doc = re.sub(r"\.0$", "", doc)
        if doc.isdigit() and len(doc) < 8:
            doc = doc.zfill(8)

        periodo = "" if get(i_periodo) is None else str(get(i_periodo)).strip()
        programa = "" if get(i_programa) is None else str(get(i_programa)).strip()
        ciclo = _to_int(get(i_ciclo), None)
        curso = "" if get(i_curso) is None else str(get(i_curso)).strip()
        nota = _to_float(get(i_nota), None)

        if not doc or not periodo or not curso or nota is None:
            continue

        out.append({
            "__row__": ridx,
            "doc": doc,
            "periodo": periodo,
            "programa": programa,
            "ciclo": ciclo,
            "curso": curso,
            "nota": nota,
        })

    return out

# ─────────────────────────────────────────────────────────────
# Catálogos CRUD
# ─────────────────────────────────────────────────────────────

class PeriodsViewSet(viewsets.ModelViewSet):
    queryset = Period.objects.all()
    serializer_class = PeriodSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "post", "patch", "delete"]

    def get_queryset(self):
        qs = super().get_queryset()
        if hasattr(Period, "start_date"):
            return qs.order_by("-start_date")
        return qs.order_by("-id")

    def list(self, request, *args, **kwargs):
        return list_items(self.serializer_class, self.get_queryset())

    @action(detail=True, methods=["post"], url_path="active")
    def set_active(self, request, pk=None):
        is_active = bool(request.data.get("is_active", False))
        p = self.get_object()
        if is_active:
            Period.objects.update(is_active=False)
        p.is_active = is_active
        p.save(update_fields=["is_active"])
        return Response({"ok": True, "id": p.id, "is_active": p.is_active})

class CampusesViewSet(viewsets.ModelViewSet):
    queryset = Campus.objects.all().order_by("name")
    serializer_class = CampusSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "post", "patch", "delete"]

    def list(self, request, *args, **kwargs):
        return list_items(self.serializer_class, self.get_queryset())

class ClassroomsViewSet(viewsets.ModelViewSet):
    queryset = Classroom.objects.select_related("campus").all()
    serializer_class = ClassroomSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "post", "patch", "delete"]

    def get_queryset(self):
        qs = super().get_queryset()
        campus_id = self.request.query_params.get("campus_id")
        if campus_id:
            qs = qs.filter(campus_id=campus_id)
        if hasattr(Classroom, "code"):
            return qs.order_by("campus__name", "code")
        return qs.order_by("campus__name", "id")

    def list(self, request, *args, **kwargs):
        return list_items(self.serializer_class, self.get_queryset())

class TeachersViewSet(viewsets.ModelViewSet):
    queryset = Teacher.objects.select_related("user").all()
    serializer_class = TeacherSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "post", "patch", "delete"]

    def get_queryset(self):
        qs = super().get_queryset()
        q = (self.request.query_params.get("q") or "").strip()
        if q:
            cond = Q()
            if hasattr(Teacher, "document"):
                cond |= Q(document__icontains=q)
            if hasattr(Teacher, "email"):
                cond |= Q(email__icontains=q)
            if hasattr(Teacher, "phone"):
                cond |= Q(phone__icontains=q)
            if hasattr(Teacher, "specialization"):
                cond |= Q(specialization__icontains=q)
            cond |= (Q(user__full_name__icontains=q) | Q(user__username__icontains=q) | Q(user__email__icontains=q))
            qs = qs.filter(cond)
        return qs.order_by("user__full_name", "user__username", "id")

    def list(self, request, *args, **kwargs):
        return list_items(self.serializer_class, self.get_queryset())

    def create(self, request, *args, **kwargs):
        teacher_role, _ = Role.objects.get_or_create(name="TEACHER")

        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        # username base: documento o primer nombre
        username_base = (data.get("document", "") or "").strip() or (data.get("full_name", "").split()[0] if (data.get("full_name") or "").strip() else "teacher")
        username = username_base
        k = 1
        while User.objects.filter(username=username).exists():
            k += 1
            username = f"{username_base}-{k}"

        # ✅ password temporal
        temp_password = get_random_string(12)

        # email fallback
        email = (data.get("email", "") or "").strip().lower()
        if not email:
            email = f"{username}@no-email.local"

        # ✅ crea user (temporal) y OBLIGA CAMBIO
        user = User.objects.create_user(
            username=username,
            password=temp_password,
            email=email,
            full_name=data.get("full_name", "") or "",
        )
        user.must_change_password = True
        user.save(update_fields=["must_change_password"])

        # ✅ asignar rol TEACHER (compat con tu esquema ACL)
        try:
            user.roles.add(teacher_role)  # si tu User tiene ManyToMany roles
        except Exception:
            UserRole.objects.get_or_create(user=user, role=teacher_role)

        # Crear Teacher
        teacher = ser.save(user=user)

        # ✅ Asignar cursos (ManyToMany)
        courses = data.get("courses", [])
        if courses is not None and hasattr(teacher, "courses"):
            teacher.courses.set(courses or [])

        out = self.get_serializer(teacher).data
        out["username"] = user.username
        out["temporary_password"] = temp_password
        out["must_change_password"] = True
        return Response(out, status=201)


UBIGEO_DEMO = {
    "15": {  # Lima
        "name": "LIMA",
        "provinces": {
            "1501": {"name": "LIMA", "districts": {"150101": "LIMA", "150114": "LA MOLINA", "150140": "SURCO", "150122": "MIRAFLORES"}},
            "1508": {"name": "HUAURA", "districts": {"150801": "HUACHO", "150802": "HUALMAY"}},
        },
    },
    "20": {  # Piura
        "name": "PIURA",
        "provinces": {
            "2001": {"name": "PIURA", "districts": {"200101": "PIURA", "200104": "CASTILLA"}},
        },
    },
}

_UBIGEO_CACHE: Optional[dict] = None

def _load_ubigeo_pe() -> dict:
    global _UBIGEO_CACHE
    if isinstance(_UBIGEO_CACHE, dict):
        return _UBIGEO_CACHE

    # Intentar rutas típicas (según dónde esté corriendo BASE_DIR)
    candidates = [
        os.path.join(settings.BASE_DIR, "catalogs", "data", "ubigeo_pe.json"),
        os.path.join(settings.BASE_DIR, "backend", "catalogs", "data", "ubigeo_pe.json"),
        os.path.join(os.getcwd(), "catalogs", "data", "ubigeo_pe.json"),
        os.path.join(os.getcwd(), "backend", "catalogs", "data", "ubigeo_pe.json"),
    ]

    last_err = None

    for path in candidates:
        try:
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)

                if isinstance(data, dict) and data:
                    _UBIGEO_CACHE = data
                    print(f"[UBIGEO] OK loaded: {path} (deps={len(data)})")
                    return data
                else:
                    last_err = f"JSON vacío o inválido en {path}"
        except Exception as e:
            last_err = f"{type(e).__name__}: {e} ({path})"

    print(f"[UBIGEO] FALLBACK DEMO. BASE_DIR={settings.BASE_DIR} CWD={os.getcwd()} ERR={last_err}")
    _UBIGEO_CACHE = UBIGEO_DEMO
    return _UBIGEO_CACHE

def _ub_name(x) -> str:
    return (x or "").strip()

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def ubigeo_departments(request):
    ub = _load_ubigeo_pe()
    out = [{"code": k, "name": _ub_name(v.get("name"))} for k, v in ub.items()]
    out.sort(key=lambda x: x["name"])
    return Response(out)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def ubigeo_provinces(request):
    ub = _load_ubigeo_pe()
    dep = (request.query_params.get("department") or "").strip()
    if dep not in ub:
        return Response([])
    provs = ub[dep].get("provinces") or {}
    out = [{"code": k, "name": _ub_name(v.get("name"))} for k, v in provs.items()]
    out.sort(key=lambda x: x["name"])
    return Response(out)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def ubigeo_districts(request):
    ub = _load_ubigeo_pe()
    dep = (request.query_params.get("department") or "").strip()
    prov = (request.query_params.get("province") or "").strip()
    if dep not in ub:
        return Response([])
    provs = ub[dep].get("provinces") or {}
    if prov not in provs:
        return Response([])
    dists = provs[prov].get("districts") or {}
    out = [{"code": k, "name": _ub_name(v)} for k, v in dists.items()]
    out.sort(key=lambda x: x["name"])
    return Response(out)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def ubigeo_search(request):
    """
    q busca por name o code. Devuelve hasta 50 coincidencias.
    """
    q = (request.query_params.get("q") or "").strip()
    if not q:
        return Response([])

    qq = _norm(q)
    ub = _load_ubigeo_pe()
    res = []

    for dep_code, dep in ub.items():
        dep_name = dep.get("name", "")
        if qq in _norm(dep_name) or qq in _norm(dep_code):
            res.append({"department": {"code": dep_code, "name": dep_name}})

        provs = dep.get("provinces") or {}
        for prov_code, prov in provs.items():
            prov_name = prov.get("name", "")
            if qq in _norm(prov_name) or qq in _norm(prov_code):
                res.append({
                    "department": {"code": dep_code, "name": dep_name},
                    "province": {"code": prov_code, "name": prov_name},
                })

            dists = prov.get("districts") or {}
            for dist_code, dist_name in dists.items():
                if qq in _norm(dist_name) or qq in _norm(dist_code):
                    res.append({
                        "department": {"code": dep_code, "name": dep_name},
                        "province": {"code": prov_code, "name": prov_name},
                        "district": {"code": dist_code, "name": dist_name},
                    })

        if len(res) >= 50:
            break

    return Response(res[:50])

# ─────────────────────────────────────────────────────────────
# Institution settings + media
# ─────────────────────────────────────────────────────────────

@api_view(["GET", "PATCH"])
@permission_classes([IsAuthenticated])
def institution_settings(request):
    obj, _ = InstitutionSetting.objects.get_or_create(pk=1)
    if request.method == "GET":
        return Response(obj.data or {})
    obj.data = {**(obj.data or {}), **(request.data or {})}
    obj.save(update_fields=["data"])
    return Response(obj.data)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def institution_media(request):
    file = request.FILES.get("file")
    kind = (request.POST.get("kind") or "").upper().strip()

    if not file or kind not in ("LOGO", "SIGNATURE"):
        return Response({"detail": "file y kind (LOGO|SIGNATURE) requeridos"}, status=400)

    # ✅ 1) crear asset
    asset = MediaAsset.objects.create(kind=kind, file=file)

    # ✅ 2) url usable (relativa)
    rel_url = asset.file.url  # ej: /media/uploads/logo.png

    # ✅ 3) guardar en InstitutionSetting.data
    inst, _ = InstitutionSetting.objects.get_or_create(pk=1)
    data = inst.data or {}

    if kind == "LOGO":
        data["logo_url"] = rel_url
    elif kind == "SIGNATURE":
        data["signature_url"] = rel_url

    inst.data = data
    inst.save(update_fields=["data"])

    # ✅ 4) también devuelve url absoluta si quieres
    abs_url = request.build_absolute_uri(rel_url)

    return Response({
        "ok": True,
        "kind": kind,
        "url": rel_url,
        "absolute_url": abs_url,
        "asset": MediaAssetSerializer(asset).data,
        "institution": inst.data,
    }, status=201)

@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def institution_media_delete(request, kind: str):
    kind = (kind or "").upper().strip()
    if kind not in ("LOGO", "SIGNATURE"):
        return Response({"detail": "kind inválido (LOGO|SIGNATURE)"}, status=400)

    inst, _ = InstitutionSetting.objects.get_or_create(pk=1)
    data = inst.data or {}

    key = "logo_url" if kind == "LOGO" else "signature_url"
    url = data.get(key)

    # ✅ limpia config
    if key in data:
        data[key] = ""
    inst.data = data
    inst.save(update_fields=["data"])

    # ✅ opcional: borrar archivo físico si url apunta a /media/...
    try:
        if url and isinstance(url, str) and "/media/" in url:
            rel = url.split("/media/")[-1]
            fpath = os.path.join(settings.MEDIA_ROOT, rel)
            if os.path.exists(fpath):
                os.remove(fpath)
    except Exception:
        pass

    return Response({"ok": True, "kind": kind, "cleared": key, "institution": inst.data})

# ─────────────────────────────────────────────────────────────
# Import templates
# ─────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def imports_template(request, type: str):
    not_ok = _require_staff_or_teacher(request)
    if not_ok:
        return not_ok

    type = (type or "").lower().strip()

    FILES = {
        "students": "students_template.xlsx",
        "grades": "grades_template.xlsx",
        "plans": "plan_estudios.xlsx",
    }

    filename = FILES.get(type)
    if not filename:
        return Response({"detail": "Tipo inválido"}, status=400)

    file_path = os.path.join(settings.BASE_DIR, "templates", "imports", filename)
    if not os.path.exists(file_path):
        raise Http404("Plantilla no encontrada")

    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ─────────────────────────────────────────────────────────────
# Import start + status
# ─────────────────────────────────────────────────────────────

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def imports_start(request, type: str):
    not_ok = _require_staff_or_teacher(request)
    if not_ok:
        return not_ok

    type = (type or "").lower().strip()
    file = request.FILES.get("file")
    mapping_raw = request.POST.get("mapping")
    mapping = {}

    if mapping_raw:
        try:
            mapping = json.loads(mapping_raw)
        except Exception:
            mapping = {}

    if not file:
        return Response({"detail": "file requerido"}, status=400)

    # ✅ Crea job con estado base
    job = _retry_db(lambda: ImportJob.objects.create(
        type=type,
        file=file,
        mapping=mapping,
        status="RUNNING",
        result={"progress": 0, "total": 0, "processed": 0, "errors": [], "imported": 0, "updated": 0}
    ))

    errors: List[dict] = []
    imported = 0
    updated = 0
    credentials: List[dict] = []

    # -------------------------
    # helpers progreso + errores
    # -------------------------
    def set_job_state(processed: int, total: int, message: str = ""):
        p = int((processed / total) * 100) if total else 0
        job.result = {
            **(job.result or {}),
            "progress": p,
            "processed": processed,
            "total": total,
            "message": message,
        }
        job.save(update_fields=["result"])

    def add_error(row, field, message):
        errors.append({"row": row, "field": field, "message": message})

    try:
        # =========================
        # PLANS
        # =========================
        if type == "plans":
            fname = (getattr(file, "name", "") or "").lower()
            if fname.endswith(".xls"):
                job.status = "FAILED"
                job.result = {**(job.result or {}), "errors": [{"row": None, "field": "file", "message": "Convierte el plan a .xlsx (openpyxl no lee .xls)."}]}
                job.save(update_fields=["status", "result"])
                return Response({"job_id": job.id}, status=400)

            if not fname.endswith(".xlsx"):
                job.status = "FAILED"
                job.result = {**(job.result or {}), "errors": [{"row": None, "field": "file", "message": "Para Plan de estudios sube un archivo .xlsx."}]}
                job.save(update_fields=["status", "result"])
                return Response({"job_id": job.id}, status=400)

            # Prelectura (solo para contar hojas/carreras)
            try:
                wb_tmp = load_workbook(file, read_only=True, keep_links=False)
                sheet_names = list(wb_tmp.sheetnames)
            except Exception as e:
                add_error(None, "file", f"No se pudo leer el .xlsx: {e}")
                sheet_names = []

            try:
                file.seek(0)
            except Exception:
                pass

            plan_rows = _read_study_plan_xlsx(file)
            total = len(plan_rows)
            set_job_state(0, total, "Leyendo plan de estudios...")

            if not plan_rows:
                add_error(None, "format", "No se detectaron filas del plan. Revisa que exista header 'AREAS/ASIGNATURAS' y 'CRED.'")

            created_careers = 0
            created_plans = 0
            created_courses = 0
            created_links = 0

            cleared_plans = set()
            done = 0

            for row in plan_rows:
                done += 1
                if done % 25 == 0:
                    set_job_state(done, total, "Procesando plan de estudios...")

                r = row.get("__row__", "?")
                career_name = (row.get("career_name") or "").strip()
                semester = row.get("semester")
                course_name = (row.get("course_name") or "").strip()
                credits = int(row.get("credits") or 0)
                hours = int(row.get("hours") or 0)

                if not career_name:
                    add_error(r, "career_name", "Carrera/Programa vacío")
                    continue
                if not semester:
                    add_error(r, "semester", "Semestre vacío o inválido")
                    continue
                if not course_name:
                    add_error(r, "course_name", "Curso vacío")
                    continue

                car = Career.objects.filter(name__iexact=career_name).first()
                if not car:
                    car = Career.objects.create(name=career_name)
                    created_careers += 1

                plan, plan_created = Plan.objects.get_or_create(
                    career=car,
                    name=f"Plan {career_name}",
                    defaults={"start_year": date.today().year, "semesters": 10},
                )
                if plan_created:
                    created_plans += 1

                if plan.id not in cleared_plans:
                    PlanCourse.objects.filter(plan=plan).delete()
                    cleared_plans.add(plan.id)

                course = Course.objects.filter(name__iexact=course_name).first()
                if not course:
                    base = _slug_code(course_name, 10)
                    code = base
                    k = 1
                    while Course.objects.filter(code=code).exists():
                        k += 1
                        code = f"{base[:8]}{k:02d}"
                    course = Course.objects.create(code=code, name=course_name, credits=max(0, credits))
                    created_courses += 1
                else:
                    if int(credits) > 0 and int(course.credits or 0) != int(credits):
                        course.credits = int(credits)
                        course.save(update_fields=["credits"])

                pc, pc_created = PlanCourse.objects.get_or_create(
                    plan=plan,
                    course=course,
                    defaults={"semester": 1, "weekly_hours": 3, "type": "MANDATORY"},
                )
                pc.semester = max(1, int(semester))
                pc.weekly_hours = max(0, int(hours)) if hours else int(pc.weekly_hours or 3)
                pc.type = pc.type or "MANDATORY"
                pc.save(update_fields=["semester", "weekly_hours", "type"])

                if pc_created:
                    created_links += 1

                imported += 1

            # Ajustar semestres reales
            for plan_id in list(cleared_plans):
                pl = Plan.objects.filter(id=plan_id).first()
                if not pl:
                    continue
                mx = (PlanCourse.objects.filter(plan_id=plan_id).aggregate(mx=models.Max("semester")).get("mx")) or 0
                if mx > 0 and int(pl.semesters or 0) != int(mx):
                    pl.semesters = int(mx)
                    pl.save(update_fields=["semesters"])

            set_job_state(total, total, "Finalizando importación de plan...")

            job.result = {
                **(job.result or {}),
                "imported": imported,
                "updated": updated,
                "errors": errors,
                "stats": {
                    "created_careers": created_careers,
                    "created_plans": created_plans,
                    "created_courses": created_courses,
                    "created_links": created_links,
                    "sheets": len(sheet_names),
                },
                "summary": {
                    "ok": imported + updated,
                    "failed": len(errors),
                    "note": "Corrige las filas con error y vuelve a importar el archivo.",
                },
            }

        # =========================
        # STUDENTS / COURSES / GRADES
        # =========================
        elif type in ("students", "courses", "grades"):
            rows = _read_rows(file, mapping)
            total = len(rows)
            set_job_state(0, total, "Leyendo archivo...")

            # ------------- STUDENTS -------------
            if type == "students":
                student_role, _ = _retry_db(lambda: Role.objects.get_or_create(name="STUDENT"))
                user_fields = {f.name for f in User._meta.fields}

                def _email_field_info():
                    try:
                        f = User._meta.get_field("email")
                        return {"exists": True, "unique": bool(getattr(f, "unique", False)), "null": bool(getattr(f, "null", False))}
                    except Exception:
                        return {"exists": False, "unique": False, "null": False}

                EMAIL_INFO = _email_field_info()

                def _safe_unique_email(username: str, email: str):
                    email = (email or "").strip().lower()
                    if email:
                        conflict = User.objects.filter(email__iexact=email).exists()
                        if not conflict:
                            return email
                        add_error(None, "email", f"Email duplicado '{email}' (se usará dummy/NULL)")

                    if EMAIL_INFO["null"]:
                        return None
                    if EMAIL_INFO["unique"]:
                        base = f"{username}@no-email.local"
                        e = base
                        k = 1
                        while User.objects.filter(email__iexact=e).exists():
                            k += 1
                            e = f"{username}-{k}@no-email.local"
                        return e
                    return ""

                def _set_name_fields(u, full_name: str):
                    if "full_name" in user_fields:
                        u.full_name = full_name
                    elif "name" in user_fields:
                        u.name = full_name

                def _ensure_user_for_student(st: Student, username: str, email: str, full_name: str, r: Any):
                    if st.user_id:
                        _retry_db(lambda: UserRole.objects.get_or_create(user_id=st.user_id, role_id=student_role.id))
                        return st.user, None

                    email_clean = (email or "").strip().lower()
                    user = User.objects.filter(username=username).first()

                    if not user and email_clean:
                        user = User.objects.filter(email__iexact=email_clean).first()

                    if user and Student.objects.filter(user_id=user.id).exists():
                        add_error(r, "user", f"user '{getattr(user,'username','')}' ya enlazado a otro estudiante")
                        return None, None

                    temp_password = None

                    if not user:
                        uname = username
                        k = 1
                        while User.objects.filter(username=uname).exists():
                            k += 1
                            uname = f"{username}-{k}"

                        user = User(username=uname, is_active=True, is_staff=False)

                        if "email" in user_fields:
                            user.email = _safe_unique_email(uname, email_clean)

                        _set_name_fields(user, full_name)

                        temp_password = get_random_string(10) + "!"
                        user.set_password(temp_password)

                        try:
                            _retry_db(lambda: user.save())
                        except IntegrityError as e:
                            if "email" in str(e).lower() and "email" in user_fields:
                                user.email = _safe_unique_email(user.username, "")
                                _retry_db(lambda: user.save())
                                add_error(r, "email", "Choque UNIQUE email (reintento ok)")
                            else:
                                raise
                    else:
                        changed = False
                        _set_name_fields(user, full_name)
                        changed = True

                        if "email" in user_fields and email_clean:
                            conflict = User.objects.filter(email__iexact=email_clean).exclude(id=user.id).exists()
                            if not conflict and (getattr(user, "email", "") or "").lower() != email_clean:
                                user.email = email_clean
                                changed = True

                        if not getattr(user, "is_active", True):
                            user.is_active = True
                            changed = True

                        if changed:
                            try:
                                _retry_db(lambda: user.save())
                            except IntegrityError as e:
                                if "email" in str(e).lower():
                                    add_error(r, "email", "Update email falló por duplicado (omitido)")
                                else:
                                    raise

                    _retry_db(lambda: UserRole.objects.get_or_create(user_id=user.id, role_id=student_role.id))

                    st.user = user
                    _retry_db(lambda: st.save(update_fields=["user"]))
                    return user, temp_password

                done = 0
                for row in rows:
                    done += 1
                    if done % 25 == 0:
                        set_job_state(done, total, "Importando alumnos...")

                    r = row.get("__row__", "?")
                    num_documento = str(row.get("num_documento", "")).strip()
                    nombres = str(row.get("nombres", "")).strip()
                    ap_pat = str(row.get("apellido_paterno", "")).strip()
                    ap_mat = str(row.get("apellido_materno", "")).strip()
                    sexo = str(row.get("sexo", "")).strip()

                    fecha_val = row.get("fecha_nac")
                    if hasattr(fecha_val, "year") and hasattr(fecha_val, "month") and hasattr(fecha_val, "day"):
                        fecha_nac = fecha_val
                    else:
                        fecha_nac = _parse_date_yyyy_mm_dd(str(fecha_val).strip())

                    if not num_documento:
                        add_error(r, "num_documento", "Num Documento es requerido")
                        continue
                    if not nombres:
                        add_error(r, "nombres", "Nombres es requerido")
                        continue

                    periodo = str(row.get("periodo", "")).strip()
                    if periodo:
                        y, t = _parse_period_code(periodo)
                        if not (y and t):
                            add_error(r, "periodo", "Periodo inválido (ej: 2026-I)")
                        else:
                            _retry_db(lambda: Period.objects.get_or_create(
                                code=periodo,
                                defaults={"year": y, "term": t, "label": periodo, "is_active": False},
                            ))

                    programa_carrera = str(row.get("programa_carrera", "")).strip()
                    plan_obj = None
                    if programa_carrera:
                        car, _ = _retry_db(lambda: Career.objects.get_or_create(name=programa_carrera))
                        plan_obj, _ = _retry_db(lambda: Plan.objects.get_or_create(
                            career=car,
                            name=f"Plan {programa_carrera}",
                            defaults={"start_year": date.today().year, "semesters": 10},
                        ))

                    st, created = _retry_db(lambda: Student.objects.get_or_create(
                        num_documento=num_documento,
                        defaults={"nombres": nombres, "apellido_paterno": ap_pat, "apellido_materno": ap_mat},
                    ))

                    st.nombres = nombres
                    st.apellido_paterno = ap_pat
                    st.apellido_materno = ap_mat
                    st.sexo = sexo
                    if fecha_nac:
                        st.fecha_nac = fecha_nac

                    st.region = str(row.get("region", "")).strip()
                    st.provincia = str(row.get("provincia", "")).strip()
                    st.distrito = str(row.get("distrito", "")).strip()
                    st.codigo_modular = str(row.get("codigo_modular", "")).strip()
                    st.nombre_institucion = str(row.get("nombre_institucion", "")).strip()
                    st.gestion = str(row.get("gestion", "")).strip()
                    st.tipo = str(row.get("tipo", "")).strip()

                    st.programa_carrera = programa_carrera
                    st.ciclo = _to_int(row.get("ciclo"), None)
                    st.turno = str(row.get("turno", "")).strip()
                    st.seccion = str(row.get("seccion", "")).strip()
                    st.periodo = periodo
                    st.lengua = str(row.get("lengua", "")).strip()
                    st.discapacidad = str(row.get("discapacidad", "")).strip()
                    st.tipo_discapacidad = str(row.get("tipo_discapacidad", "")).strip()

                    st.email = str(row.get("email", "") or "").strip().lower()
                    st.celular = str(row.get("celular", "") or "").strip()
                    st.plan = plan_obj

                    username = num_documento
                    full_name = f"{nombres} {ap_pat} {ap_mat}".strip()
                    user, temp_password = _ensure_user_for_student(st, username, st.email, full_name, r)
                    _retry_db(lambda: st.save())

                    if temp_password and user:
                        credentials.append({"row": r, "num_documento": num_documento, "username": getattr(user, "username", username), "password": temp_password})

                    if created:
                        imported += 1
                    else:
                        updated += 1

                set_job_state(total, total, "Finalizando alumnos...")

            # ------------- COURSES -------------
            elif type == "courses":
                done = 0
                for row in rows:
                    done += 1
                    if done % 25 == 0:
                        set_job_state(done, total, "Importando cursos...")

                    r = row.get("__row__", "?")
                    code = str(row.get("code", "")).strip()
                    name = str(row.get("name", "")).strip()

                    if not code:
                        add_error(r, "code", "code requerido")
                        continue
                    if not name:
                        add_error(r, "name", "name requerido")
                        continue

                    credits = _to_int(row.get("credits"), None)
                    hours = _to_int(row.get("hours", None))

                    course, _ = Course.objects.get_or_create(code=code, defaults={"name": name, "credits": max(0, credits or 0)})
                    course.name = name
                    if credits is not None:
                        course.credits = max(0, int(credits))
                    course.save()

                    plan_id = _to_int(row.get("plan_id", None))
                    semester = _to_int(row.get("semester", None))
                    ctype = str(row.get("type", "")).strip().upper()

                    if plan_id and semester:
                        plan = Plan.objects.filter(id=plan_id).first()
                        if not plan:
                            add_error(r, "plan_id", f"plan_id {plan_id} no existe")
                        else:
                            if ctype in ("OBLIGATORIO", "MANDATORY", "M"):
                                type_db = "MANDATORY"
                            elif ctype in ("ELECTIVO", "ELECTIVE", "E"):
                                type_db = "ELECTIVE"
                            else:
                                type_db = "MANDATORY"

                            pc, pc_created = PlanCourse.objects.get_or_create(
                                plan=plan,
                                course=course,
                                defaults={"semester": max(1, int(semester)), "weekly_hours": max(1, int(hours or 3)), "type": type_db},
                            )
                            if not pc_created:
                                pc.semester = max(1, int(semester))
                                pc.weekly_hours = max(1, int(hours or pc.weekly_hours or 3))
                                pc.type = type_db
                                pc.save()

                    imported += 1

                set_job_state(total, total, "Finalizando cursos...")

            # ------------- GRADES -------------
            elif type == "grades":
                if AcademicGradeRecord is None:
                    job.status = "FAILED"
                    job.result = {**(job.result or {}), "errors": [{"row": None, "field": "model", "message": "AcademicGradeRecord no existe. Agrégalo y migra."}]}
                    job.save(update_fields=["status", "result"])
                    return Response({"job_id": job.id}, status=400)

                cal_rows = _read_calificaciones_xlsx(file)
                total = len(cal_rows)
                set_job_state(0, total, "Leyendo calificaciones...")

                done = 0
                for row in cal_rows:
                    done += 1
                    if done % 25 == 0:
                        set_job_state(done, total, "Importando notas...")

                    r = row.get("__row__", "?")
                    doc = row.get("doc")
                    term = row.get("periodo")
                    ciclo = row.get("ciclo")
                    course_name = row.get("curso")
                    final = row.get("nota")

                    if not doc:
                        add_error(r, "student_document", "Documento del alumno vacío")
                        continue
                    if not term:
                        add_error(r, "term", "Periodo vacío (ej: 2026-I)")
                        continue
                    if not course_name:
                        add_error(r, "course", "Curso vacío")
                        continue
                    if final is None:
                        add_error(r, "final_grade", "Nota final vacía")
                        continue

                    st = Student.objects.filter(num_documento=doc).first()
                    if not st:
                        add_error(r, "student_document", f"No existe alumno con documento {doc}")
                        continue

                    y, t = _parse_period_code(term)
                    if y and t:
                        Period.objects.get_or_create(code=term, defaults={"year": y, "term": t, "label": term, "is_active": False})
                    else:
                        add_error(r, "term", f"Periodo inválido '{term}' (ej: 2026-I)")

                    course = Course.objects.filter(name__iexact=course_name).first()
                    if not course:
                        base = _slug_code(course_name, 10)
                        code = base
                        k = 1
                        while Course.objects.filter(code=code).exists():
                            k += 1
                            code = f"{base[:8]}{k:02d}"
                        course = Course.objects.create(code=code, name=course_name, credits=0)

                    if st.plan_id and ciclo:
                        PlanCourse.objects.get_or_create(
                            plan_id=st.plan_id,
                            course=course,
                            defaults={"semester": max(1, int(ciclo)), "weekly_hours": 3, "type": "MANDATORY"},
                        )

                    rec, created = AcademicGradeRecord.objects.get_or_create(
                        student=st,
                        course=course,
                        term=term,
                        defaults={"final_grade": float(final), "components": {}},
                    )
                    if not created:
                        rec.final_grade = float(final)
                        rec.save(update_fields=["final_grade"])

                    imported += 1

                set_job_state(total, total, "Finalizando notas...")

        else:
            job.status = "FAILED"
            job.result = {**(job.result or {}), "errors": [{"row": None, "field": "type", "message": "Tipo inválido"}]}
            job.save(update_fields=["status", "result"])
            return Response({"job_id": job.id}, status=400)

        # ✅ Estado final (si hubo errores, igual dejamos resumen)
        job.status = "COMPLETED" if len(errors) == 0 else "FAILED"
        job.result = {
            **(job.result or {}),
            "imported": imported,
            "updated": updated,
            "errors": errors,
            "credentials": credentials[:300],
            "summary": {
                "ok": imported + updated,
                "failed": len(errors),
                "note": "Corrige los errores por fila y vuelve a importar el archivo.",
            }
        }
        job.save(update_fields=["status", "result"])
        return Response({"job_id": job.id})

    except Exception as e:
        job.status = "FAILED"
        job.result = {
            **(job.result or {}),
            "errors": errors + [{"row": None, "field": "exception", "message": str(e)}],
            "imported": imported,
            "updated": updated,
        }
        try:
            job.save(update_fields=["status", "result"])
        except Exception:
            pass
        return Response({"job_id": job.id, "detail": "Import failed", "error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def imports_status(request, jobId: int):
    try:
        job = ImportJob.objects.get(pk=jobId)
    except ImportJob.DoesNotExist:
        return Response({"detail": "job not found"}, status=404)

    result = job.result if isinstance(job.result, dict) else {}
    return Response({
        "id": job.id,
        "state": job.status,
        "progress": int(result.get("progress") or 0),
        "processed": int(result.get("processed") or 0),
        "total": int(result.get("total") or 0),
        "message": result.get("message") or "",
        "errors": result.get("errors") or [],
        "imported": result.get("imported", 0),
        "updated": result.get("updated", 0),
        "credentials": result.get("credentials") or [],
        "summary": result.get("summary") or {},
        "stats": result.get("stats") or {},
    })

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def imports_status(request, jobId: int):
    try:
        job = ImportJob.objects.get(pk=jobId)
    except ImportJob.DoesNotExist:
        return Response({"detail": "job not found"}, status=404)

    result = job.result if isinstance(job.result, dict) else {}
    return Response({
        "id": job.id,
        "state": job.status,
        "progress": 100 if job.status in ("COMPLETED", "FAILED") else 0,
        "errors": result.get("errors") or [],
        "imported": result.get("imported", 0),
        "updated": result.get("updated", 0),
        "credentials": result.get("credentials") or [],
        "stats": result.get("stats") or {},
    })

# ─────────────────────────────────────────────────────────────
# Backups / Export dataset
# ─────────────────────────────────────────────────────────────

def _zip_add_folder(zf: zipfile.ZipFile, folder_path: str, arc_prefix: str):
    folder_path = os.path.abspath(folder_path)
    if not os.path.exists(folder_path):
        return
    for root, _, files in os.walk(folder_path):
        for f in files:
            fp = os.path.join(root, f)
            rel = os.path.relpath(fp, folder_path).replace("\\", "/")
            zf.write(fp, f"{arc_prefix}/{rel}")

def _dumpdata_json_bytes():
    buf = io.StringIO()
    call_command("dumpdata", "--natural-foreign", "--natural-primary", "--indent", "2", stdout=buf)
    return buf.getvalue().encode("utf-8")

def _try_add_sqlite_db(zf: zipfile.ZipFile):
    try:
        db = settings.DATABASES.get("default", {})
        if db.get("ENGINE") == "django.db.backends.sqlite3":
            name = db.get("NAME")
            if name and os.path.exists(name):
                zf.write(str(name), "db/db.sqlite3")
    except Exception:
        pass

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def backups_collection(request):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    if request.method == "GET":
        qs = BackupExport.objects.all().order_by("-created_at")
        return Response({"items": BackupExportSerializer(qs, many=True).data})

    scope = (request.data.get("scope") or "FULL").upper().strip()
    if scope not in ("FULL", "DATA_ONLY", "FILES_ONLY"):
        return Response({"detail": "scope inválido (FULL|DATA_ONLY|FILES_ONLY)"}, status=400)

    obj = BackupExport.objects.create(scope=scope)

    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_name = f"backup_{scope.lower()}_{now}.zip"

    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        if scope in ("FULL", "DATA_ONLY"):
            zf.writestr("data/dumpdata.json", _dumpdata_json_bytes())
            _try_add_sqlite_db(zf)

        if scope in ("FULL", "FILES_ONLY"):
            media_root = getattr(settings, "MEDIA_ROOT", None)
            if media_root:
                _zip_add_folder(zf, str(media_root), "media")

        zf.writestr("meta/info.txt", f"scope={scope}\ncreated_at={now}\n")

    zbuf.seek(0)
    obj.file.save(zip_name, ContentFile(zbuf.getvalue()))
    obj.save(update_fields=["file"])

    return Response({
        "id": obj.id,
        "scope": obj.scope,
        "file_url": obj.file.url if obj.file else None,
    }, status=201)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def backup_download(request, id: int):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    try:
        b = BackupExport.objects.get(pk=id)
    except BackupExport.DoesNotExist:
        raise Http404

    if not b.file:
        raise Http404

    return FileResponse(b.file.open("rb"), as_attachment=True, filename=os.path.basename(b.file.name))

@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def backup_delete(request, id: int):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    try:
        b = BackupExport.objects.get(pk=id)
    except BackupExport.DoesNotExist:
        raise Http404

    # ✅ Borra archivo físico (MEDIA) si existe
    try:
        if b.file:
            b.file.delete(save=False)
    except Exception:
        # si el archivo ya no existe, igual borramos el registro
        pass

    b.delete()
    return Response({"ok": True, "deleted_id": id})

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def backups_cleanup(request):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    days = int(request.data.get("days") or 30)
    only_datasets = bool(request.data.get("only_datasets") or False)

    cutoff = timezone.now() - timedelta(days=days)

    qs = BackupExport.objects.filter(created_at__lt=cutoff)
    if only_datasets:
        qs = qs.filter(scope__startswith="DATASET_")

    deleted = 0
    for b in qs:
        try:
            if b.file:
                b.file.delete(save=False)
        except Exception:
            pass
        b.delete()
        deleted += 1

    return Response({"ok": True, "deleted": deleted, "days": days, "only_datasets": only_datasets})

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def export_dataset(request):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    dataset = (request.data.get("dataset") or "DATA").upper().strip()
    now = datetime.now().strftime("%Y%m%d_%H%M%S")

    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        if dataset == "STUDENTS":
            rows = list(Student.objects.all().values(
                "num_documento","nombres","apellido_paterno","apellido_materno",
                "sexo","fecha_nac",
                "region","provincia","distrito",
                "codigo_modular","nombre_institucion","gestion","tipo",
                "programa_carrera","ciclo","turno","seccion","periodo",
                "lengua","discapacidad","tipo_discapacidad",
                "email","celular","plan_id","user_id",
            ))
            headers = [
                "num_documento","nombres","apellido_paterno","apellido_materno",
                "sexo","fecha_nac",
                "region","provincia","distrito",
                "codigo_modular","nombre_institucion","gestion","tipo",
                "programa_carrera","ciclo","turno","seccion","periodo",
                "lengua","discapacidad","tipo_discapacidad",
                "email","celular","plan_id","user_id",
            ]
            zf.writestr(f"students_{now}.csv", _csv_bytes(rows, headers))

        elif dataset == "COURSES":
            rows = list(Course.objects.all().values("code","name","credits"))
            headers = ["code","name","credits"]
            zf.writestr(f"courses_{now}.csv", _csv_bytes(rows, headers))

            pc_rows = list(PlanCourse.objects.all().values("plan_id","course_id","semester","weekly_hours","type"))
            pc_headers = ["plan_id","course_id","semester","weekly_hours","type"]
            zf.writestr(f"plan_courses_{now}.csv", _csv_bytes(pc_rows, pc_headers))

        elif dataset == "GRADES":
            if AcademicGradeRecord is None:
                return Response({"detail": "AcademicGradeRecord no existe. Agrégalo y migra."}, status=400)

            rows = list(AcademicGradeRecord.objects.select_related("student","course").values(
                "student__num_documento","course__code","term","final_grade","components"
            ))
            headers = ["student__num_documento","course__code","term","final_grade","components"]
            zf.writestr(f"grades_{now}.csv", _csv_bytes(rows, headers))

        elif dataset == "CATALOGS":
            p_rows = list(Period.objects.all().values("code","year","term","start_date","end_date","is_active","label"))
            p_headers = ["code","year","term","start_date","end_date","is_active","label"]
            zf.writestr(f"catalog_periods_{now}.csv", _csv_bytes(p_rows, p_headers))

            c_rows = list(Campus.objects.all().values("code","name","address"))
            c_headers = ["code","name","address"]
            zf.writestr(f"catalog_campuses_{now}.csv", _csv_bytes(c_rows, c_headers))

            a_rows = list(Classroom.objects.all().values("campus_id","code","name","capacity"))
            a_headers = ["campus_id","code","name","capacity"]
            zf.writestr(f"catalog_classrooms_{now}.csv", _csv_bytes(a_rows, a_headers))

            t_rows = list(Teacher.objects.all().values("document","full_name","email","phone","specialization"))
            t_headers = ["document","full_name","email","phone","specialization"]
            zf.writestr(f"catalog_teachers_{now}.csv", _csv_bytes(t_rows, t_headers))

            ac_rows = list(Career.objects.all().values("id","name"))
            ac_headers = ["id","name"]
            zf.writestr(f"academic_careers_{now}.csv", _csv_bytes(ac_rows, ac_headers))

            pl_rows = list(Plan.objects.all().values("id","career_id","name","start_year","semesters"))
            pl_headers = ["id","career_id","name","start_year","semesters"]
            zf.writestr(f"academic_plans_{now}.csv", _csv_bytes(pl_rows, pl_headers))

        else:
            return Response({"detail": "dataset inválido"}, status=400)

        zf.writestr("meta/info.txt", f"dataset={dataset}\ncreated_at={now}\n")

    zbuf.seek(0)
    obj = BackupExport.objects.create(scope=f"DATASET_{dataset}")
    filename = f"export_{dataset.lower()}_{now}.zip"
    obj.file.save(filename, ContentFile(zbuf.getvalue()))
    obj.save(update_fields=["file"])

    return Response({
        "ok": True,
        "dataset": dataset,
        "backup_id": obj.id,
        "download_url": f"/catalogs/exports/backups/{obj.id}/download",
    })