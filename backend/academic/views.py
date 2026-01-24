from datetime import datetime
from io import BytesIO
import random
import csv
import base64
import os
import re
import unicodedata
from django.conf import settings
from openpyxl import load_workbook, Workbook
from django.template.loader import render_to_string
from academic.pdf_render import html_to_pdf_bytes
from django.db import transaction, IntegrityError # ✅ IntegrityError importado
from django.db.models import Q, Count
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.core.exceptions import FieldError
from django.contrib.auth import get_user_model
from django.utils import timezone
from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework_simplejwt.authentication import JWTAuthentication
from students.models import Student as StudentProfile
from academic.models import AcademicGradeRecord, Career
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from pypdf import PdfReader, PdfWriter
from acl.models import Role, UserRole
import requests
from .models import (
    Plan, PlanCourse, Course, CoursePrereq,
    Teacher, Classroom,
    Section, SectionScheduleSlot,
    AttendanceSession, AttendanceRow,
    Syllabus, EvaluationConfig,
    AcademicProcess, ProcessFile,
    SectionGrades,InstitutionSettings,Course
)
from .serializers import (
    PlanSerializer, PlanCreateSerializer,
    PlanCourseOutSerializer, PlanCourseCreateSerializer,
    ClassroomSerializer,
    SectionOutSerializer, SectionCreateUpdateSerializer,
    AttendanceSessionSerializer,
)
import logging  # Add this at the top if not already present

logger = logging.getLogger(__name__)
def url_to_data_uri(url: str) -> str:
    if not url:
        return ""
    try:
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        ctype = r.headers.get("content-type") or "image/png"
        b64 = base64.b64encode(r.content).decode("ascii")
        return f"data:{ctype};base64,{b64}"
    except Exception:
        return ""

def _build_reporte_periodo_ctx(request, st: StudentProfile, pq: str) -> tuple[dict, str]:
    pq = _norm_term(pq)

    # 1) Trae registros del periodo directamente (fuente de verdad)
    recs = (
        AcademicGradeRecord.objects
        .select_related("course")
        .filter(student=st)
    )
    recs = [r for r in recs if _norm_term(getattr(r, "term", "") or "") == pq]

    if not recs:
        return {}, "No hay registros para el periodo"

    # 2) Elige el "mejor" registro por curso (si hay repetidos)
    best_by_course = {}
    for r in recs:
        cid = r.course_id
        try:
            g = None if r.final_grade is None else float(r.final_grade)
        except Exception:
            g = None

        prev = best_by_course.get(cid)
        prev_g = None
        if prev is not None:
            try:
                prev_g = None if prev.final_grade is None else float(prev.final_grade)
            except Exception:
                prev_g = None

        if prev is None or ((g is not None) and (prev_g is None or g > prev_g)):
            best_by_course[cid] = r

    recs = list(best_by_course.values())
    recs.sort(key=lambda r: (getattr(r.course, "code", "") or "", getattr(r.course, "name", "") or ""))

    # 3) Construye filas + calcula ponderado
    rows = []
    sum_points = 0.0
    sum_credits = 0
    simple_sum = 0.0
    simple_count = 0

    for i, r in enumerate(recs, start=1):
        course = r.course
        course_name = getattr(course, "name", "") or ""

        grade = _safe_float(getattr(r, "final_grade", None))
        credits = int(getattr(course, "credits", 0) or 0)  # ✅ créditos reales del curso

        # status (si tu record guarda estado real, úsalo)
        status_text = ""
        for k in ("status", "state", "estado", "observacion", "observation"):
            if hasattr(r, k):
                status_text = (getattr(r, k) or "").strip().upper()
                break
        if status_text not in ("LOGRADO", "EN PROCESO"):
            status_text = "LOGRADO" if (grade is not None and grade >= 11) else "EN PROCESO"

        points = (grade * credits) if (grade is not None and credits > 0) else 0

        rows.append({
            "n": i,
            "course_name": course_name,
            "status_text": status_text,
            "grade": _fmt_grade(grade),
            "credits": credits,
            "points": _fmt_grade(points),
        })

        if grade is not None:
            simple_sum += float(grade)
            simple_count += 1

        if grade is not None and credits > 0:
            sum_points += float(points)
            sum_credits += credits

    # 4) Promedio: ponderado si hay créditos, si no hay créditos => simple (para no dejar vacío)
    if sum_credits > 0:
        weighted_avg = round(sum_points / sum_credits, 2)
    else:
        weighted_avg = round(simple_sum / simple_count, 2) if simple_count > 0 else ""

    logo_data, sig_data = _get_institution_media_datauris(request)

    institution_name = 'I.E.S.P.P "GUSTAVO ALLENDE LLAVERIA"'
    try:
        inst = InstitutionSettings.objects.get(id=1)
        if (inst.name or "").strip():
            institution_name = inst.name.strip()
    except Exception:
        pass

    ctx = {
        "institution_name": institution_name,
        "academic_period": pq.upper(),
        "program_name": (st.programa_carrera or "EDUCACIÓN INICIAL (RVM N° 163-2019-MINEDU)"),
        "cycle_section": getattr(st, "ciclo_seccion", "") or 'I - "A"',
        "student_name": f"{st.apellido_paterno} {st.apellido_materno} {st.nombres}".strip(),
        "shift": (getattr(st, "turno", "") or "MAÑANA").upper(),
        "enrollment_code": (getattr(st, "codigo_matricula", "") or st.num_documento or "N/A"),
        "modality": (getattr(st, "modalidad", "") or "PRESENCIAL").upper(),
        "rows": rows,
        "weighted_avg": f"{weighted_avg:.2f}" if weighted_avg != "" else "",
        "logo_url": logo_data,
        "signature_url": sig_data,
    }
    return ctx, ""

def _render_reporte_periodo_pdf(request, st: StudentProfile, pq: str) -> bytes:
    ctx, err = _build_reporte_periodo_ctx(request, st, pq)
    if err:
        raise ValueError(err)
    
    html = render_to_string("kardex/reporte_calificaciones.html", ctx)
    return html_to_pdf_bytes(html)

def _abs_media_url(request, maybe_url: str) -> str:
    if not maybe_url:
        return ""
    u = str(maybe_url).strip()
    if not u:
        return ""
    if u.startswith("http://") or u.startswith("https://"):
        return u
    if not u.startswith("/"):
        u = "/" + u
    try:
        return request.build_absolute_uri(u)
    except Exception:
        return u
def _safe_float(v):
    try:
        if v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None
        return float(s)
    except Exception:
        return None

def _fmt_grade(g):
    if g is None:
        return ""
    try:
        gf = float(g)
        if gf.is_integer():
            return str(int(gf))
        return f"{gf:.2f}".rstrip("0").rstrip(".")
    except Exception:
        return ""


def _status_text_from_record(rec):
    """
    IMPORTANTE:
    En tu foto sale 'EN PROCESO' con 14, así que NO es por nota.
    Si tu modelo guarda estado real (status/estado/observacion), úsalo.
    Si no existe, cae a regla simple (>=11 LOGRADO).
    """
    for k in ("status", "state", "estado", "observacion", "observation"):
        if hasattr(rec, k):
            v = (getattr(rec, k) or "").strip().upper()
            if v in ("LOGRADO", "EN PROCESO"):
                return v

    g = _safe_float(getattr(rec, "final_grade", None))
    if g is None:
        return "EN PROCESO"
    return "LOGRADO" if g >= 11 else "EN PROCESO"

def _get_institution_media_datauris(request):
    inst, _ = InstitutionSettings.objects.get_or_create(id=1)
    logo_abs = _abs_media_url(request, inst.logo_url)
    sig_abs = _abs_media_url(request, inst.signature_url)
    return url_to_data_uri(logo_abs), url_to_data_uri(sig_abs)

# ───────────────────────── Helpers ─────────────────────────
def ok(data=None, **extra):
    if data is None:
        data = {}
    data.update(extra)
    return Response(data)
DAY_TO_INT = {"MON": 1, "TUE": 2, "WED": 3, "THU": 4, "FRI": 5, "SAT": 6, "SUN": 7}
INT_TO_DAY = {v: k for k, v in DAY_TO_INT.items()}
def _get_full_name(u):
    if not u:
        return ""
    if hasattr(u, "get_full_name"):
        try:
            fn = (u.get_full_name() or "").strip()
            if fn:
                return fn
        except Exception:
            pass
    for attr in ("full_name", "name"):
        if hasattr(u, attr):
            v = (getattr(u, attr) or "").strip()
            if v:
                return v
    first = (getattr(u, "first_name", "") or "").strip()
    last = (getattr(u, "last_name", "") or "").strip()
    if first or last:
        return f"{first} {last}".strip()
    return (getattr(u, "username", "") or getattr(u, "email", "") or f"User {getattr(u,'id','')}").strip()
# ───────────────────────── ACL-aware role helpers ─────────────────────────
def list_users_by_role_names(role_names):
    User = get_user_model()
    role_ids = list(Role.objects.filter(name__in=role_names).values_list("id", flat=True))
    if not role_ids:
        return User.objects.none()
    user_ids = (
        UserRole.objects
        .filter(role_id__in=role_ids)
        .values_list("user_id", flat=True)
        .distinct()
    )
    return User.objects.filter(id__in=user_ids, is_active=True)
def list_teacher_users_qs():
    teacher_names = ["TEACHER", "DOCENTE", "PROFESOR"]
    return list_users_by_role_names(teacher_names)
def list_student_users_qs():
    student_names = ["STUDENT", "ALUMNO", "ESTUDIANTE"]
    return list_users_by_role_names(student_names)
def user_has_any_role(user, names):
    if not user:
        return False
    if getattr(user, "is_superuser", False):
        return True
    try:
        return UserRole.objects.filter(user=user, role__name__in=names).exists()
    except Exception:
        return False
def resolve_teacher(teacher_id):
    if not teacher_id:
        return None
    tid = int(teacher_id)
    t = Teacher.objects.select_related("user").filter(user_id=tid).first()
    if t:
        return t
    t = Teacher.objects.select_related("user").filter(id=tid).first()
    if t:
        return t
    User = get_user_model()
    u = get_object_or_404(User, id=tid)
    t, _ = Teacher.objects.get_or_create(user=u)
    return t
def resolve_classroom(room_id, code=None, capacity=None):
    if not room_id:
        return None
    defaults = {
        "code": (code or f"AULA-{room_id}"),
        "capacity": int(capacity) if capacity else 999,
    }
    room, _ = Classroom.objects.get_or_create(id=int(room_id), defaults=defaults)
    changed = False
    if code and room.code != code:
        room.code = code
        changed = True
    if capacity is not None:
        try:
            cap = int(capacity)
            if cap > 0 and room.capacity != cap:
                room.capacity = cap
                changed = True
        except Exception:
            pass
    if changed:
        room.save()
    return room
def count_teachers():
    qs = list_teacher_users_qs()
    if qs.exists():
        return qs.count()
    return Teacher.objects.count()
def _slots_for_section(section: Section):
    out = []
    for s in section.schedule_slots.all().order_by("weekday", "start"):
        out.append({
            "day": INT_TO_DAY.get(int(s.weekday), str(s.weekday)),
            "start": str(s.start)[:5],
            "end": str(s.end)[:5],
        })
    return out
def _overlaps(a_start, a_end, b_start, b_end):
    return a_end > b_start and b_end > a_start
def _detect_schedule_conflicts(sections):
    events = []
    for sec in sections:
        for sl in sec.schedule_slots.all():
            events.append((int(sl.weekday), sl.start, sl.end, sec.id))
    conflicts = []
    by_day = {}
    for wd, st, en, sid in events:
        by_day.setdefault(wd, []).append((st, en, sid))
    for wd, items in by_day.items():
        items.sort(key=lambda x: x[0])
        for i in range(len(items) - 1):
            st1, en1, s1 = items[i]
            st2, en2, s2 = items[i + 1]
            if s1 == s2:
                continue
            if _overlaps(st1, en1, st2, en2):
                conflicts.append({
                    "type": "OVERLAP",
                    "weekday": wd,
                    "a": s1,
                    "b": s2,
                    "message": f"Choque de horario (día {INT_TO_DAY.get(wd, wd)}) entre secciones {s1} y {s2}"
                })
    return conflicts
def _dummy_pdf_response(filename="documento.pdf"):
    buf = BytesIO(b"%PDF-1.4\n% Dummy PDF\n1 0 obj\n<<>>\nendobj\ntrailer\n<<>>\n%%EOF\n")
    return FileResponse(buf, as_attachment=True, filename=filename)
# =========================
# PDF Template (Kardex)
# =========================
def _norm_key(s: str) -> str:
    return (s or "").strip().lower()
def _pick_kardex_template(career_name: str) -> str:
    c = _norm_key(career_name)
    if "inicial" in c:
        return "inicial.pdf"
    if "primaria" in c:
        return "primaria.pdf"
    if "comunic" in c:
        return "comunicacion.pdf"
    if "fisic" in c:
        return "educacion_fisica.pdf"
    return "inicial.pdf"
KARDEX_POS = {
    "name": (165, 695),
    "code": (420, 678),
    "nota_x": 300,
    "row_y_start": 604,
    "row_step": 18.,
    "rows_per_page": 30,
}
def _draw_text(c, x, y, text, size=10, bold=False):
    c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
    c.drawString(x, y, "" if text is None else str(text))
def _make_overlay_pdf(num_pages: int, draw_fn):
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    for page_i in range(num_pages):
        draw_fn(c, page_i)
        c.showPage()
    c.save()
    buf.seek(0)
    return PdfReader(buf)
def _merge_overlay(template_pdf_path: str, overlay_reader: PdfReader) -> bytes:
    tpl = PdfReader(template_pdf_path)
    out = PdfWriter()
    n = min(len(tpl.pages), len(overlay_reader.pages))
    for i in range(n):
        base = tpl.pages[i]
        base.merge_page(overlay_reader.pages[i])
        out.add_page(base)
    for i in range(n, len(tpl.pages)):
        out.add_page(tpl.pages[i])
    bio = BytesIO()
    out.write(bio)
    bio.seek(0)
    return bio.getvalue()
# ─────────────────────── Available courses ───────────────────────
class AvailableCoursesView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        plan_id = request.query_params.get("plan_id")
        semester = request.query_params.get("semester")
        q = (request.query_params.get("q") or "").strip()
        academic_period = (request.query_params.get("academic_period") or "2025-I").strip()
        qs = PlanCourse.objects.select_related("plan", "course").all()
        if plan_id:
            try:
                qs = qs.filter(plan_id=int(plan_id))
            except Exception:
                return ok(courses=[])
        if semester:
            try:
                qs = qs.filter(semester=int(semester))
            except Exception:
                return ok(courses=[])
        if q:
            qs = qs.filter(Q(course__code__icontains=q) | Q(course__name__icontains=q))
        qs = qs.order_by("course__code")
        plan_course_ids = list(qs.values_list("id", flat=True))
        sections = (
            Section.objects
            .select_related("plan_course__course")
            .prefetch_related("schedule_slots")
            .filter(plan_course_id__in=plan_course_ids, period=academic_period)
            .order_by("plan_course_id", "label", "id")
        )
        secs_by_pc = {}
        for s in sections:
            secs_by_pc.setdefault(s.plan_course_id, []).append(s)
        def slots_to_str(sec):
            slots = []
            for sl in sec.schedule_slots.all().order_by("weekday", "start"):
                day = INT_TO_DAY.get(int(sl.weekday), str(sl.weekday))
                slots.append(f"{day} {str(sl.start)[:5]}-{str(sl.end)[:5]}")
            return ", ".join(slots)
        courses = []
        for pc in qs:
            secs = secs_by_pc.get(pc.id, [])
            parts = []
            for sec in secs[:3]:
                sstr = slots_to_str(sec)
                if sstr:
                    parts.append(f"{sec.label}: {sstr}")
            schedule_str = " | ".join(parts) if parts else ""
            courses.append({
                "id": pc.id,
                "code": pc.course.code,
                "name": pc.course.name,
                "credits": pc.course.credits,
                "schedule": schedule_str,
            })
        return ok(courses=courses)
# ─────────────────────── Enrollment endpoints ───────────────────────
class EnrollmentValidateView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        body = request.data or {}
        academic_period = (body.get("academic_period") or "2025-I").strip()
        course_ids = body.get("course_ids") or []
        if not isinstance(course_ids, list) or not course_ids:
            return Response({"detail": "course_ids debe ser una lista no vacía"}, status=400)
        try:
            course_ids = [int(x) for x in course_ids]
        except Exception:
            return Response({"detail": "course_ids inválidos"}, status=400)
        sections_qs = (
            Section.objects
            .select_related("plan_course__course", "teacher__user")
            .prefetch_related("schedule_slots")
            .filter(plan_course_id__in=course_ids, period=academic_period)
            .order_by("plan_course_id", "label", "id")
        )
        sections = list(sections_qs)
        chosen = {}
        for s in sections:
            if s.plan_course_id not in chosen:
                chosen[s.plan_course_id] = s
        missing = [cid for cid in course_ids if cid not in chosen]
        if missing:
            return Response({
                "errors": [f"No hay secciones disponibles en {academic_period} para course_id(s): {missing}"],
                "warnings": [],
                "suggestions": [],
                "schedule_conflicts": [],
            }, status=409)
        chosen_sections = list(chosen.values())
        conflicts = _detect_schedule_conflicts(chosen_sections)
        if conflicts:
            return Response({
                "errors": ["Se detectaron choques de horario entre cursos seleccionados."],
                "warnings": [],
                "suggestions": [],
                "schedule_conflicts": conflicts,
            }, status=409)
        return ok(warnings=[])
class EnrollmentSuggestionsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        body = request.data or {}
        academic_period = (body.get("academic_period") or "2025-I").strip()
        course_ids = body.get("course_ids") or []
        if not isinstance(course_ids, list) or not course_ids:
            return ok(suggestions=[])
        try:
            course_ids = [int(x) for x in course_ids]
        except Exception:
            return ok(suggestions=[])
        all_secs = list(
            Section.objects
            .select_related("plan_course__course", "teacher__user")
            .prefetch_related("schedule_slots")
            .filter(plan_course_id__in=course_ids, period=academic_period)
            .order_by("plan_course_id", "label", "id")
        )
        chosen = {}
        by_course = {}
        for s in all_secs:
            by_course.setdefault(s.plan_course_id, []).append(s)
            if s.plan_course_id not in chosen:
                chosen[s.plan_course_id] = s
        chosen_sections = list(chosen.values())
        base_conflicts = _detect_schedule_conflicts(chosen_sections)
        if not base_conflicts:
            return ok(suggestions=[])
        conflict_course_ids = set()
        for c in base_conflicts:
            a = next((s for s in chosen_sections if s.id == c["a"]), None)
            b = next((s for s in chosen_sections if s.id == c["b"]), None)
            if a:
                conflict_course_ids.add(a.plan_course_id)
            if b:
                conflict_course_ids.add(b.plan_course_id)
        suggestions = []
        for pc_id in conflict_course_ids:
            current = chosen.get(pc_id)
            candidates = [s for s in by_course.get(pc_id, []) if (not current or s.id != current.id)]
            others = [s for s in chosen_sections if s.plan_course_id != pc_id]
            for cand in candidates:
                test = others + [cand]
                if not _detect_schedule_conflicts(test):
                    pc = cand.plan_course
                    crs = pc.course
                    teacher_name = _get_full_name(getattr(cand.teacher, "user", None)) if cand.teacher else ""
                    slots = _slots_for_section(cand)
                    suggestions.append({
                        "course_id": pc.id,
                        "course_code": crs.code,
                        "course_name": crs.name,
                        "credits": crs.credits,
                        "section_code": cand.label,
                        "teacher_name": teacher_name,
                        "slots": slots,
                        "capacity": cand.capacity,
                        "available": cand.capacity,
                    })
                    break
        return ok(suggestions=suggestions)
class EnrollmentCommitView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        body = request.data or {}
        academic_period = (body.get("academic_period") or "2025-I").strip()
        course_ids = body.get("course_ids") or []
        if not isinstance(course_ids, list) or not course_ids:
            return Response({"detail": "course_ids debe ser lista no vacía"}, status=400)
        validate_view = EnrollmentValidateView()
        validate_view.request = request
        resp = validate_view.post(request)
        if resp.status_code == 409:
            return resp
        if resp.status_code != 200:
            return resp
        enrollment_id = int(datetime.now().timestamp() * 1000) + random.randint(1, 999)
        return ok(enrollment_id=enrollment_id, academic_period=academic_period)
class EnrollmentCertificateView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request, enrollment_id: int):
        return ok(success=True, downloadUrl=f"/api/enrollments/{enrollment_id}/certificate/pdf",
                  download_url=f"/api/enrollments/{enrollment_id}/certificate/pdf")
    def get(self, request, enrollment_id: int):
        return ok(success=True, downloadUrl=f"/api/enrollments/{enrollment_id}/certificate/pdf",
                  download_url=f"/api/enrollments/{enrollment_id}/certificate/pdf")
class EnrollmentCertificatePDFView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, enrollment_id: int):
        return _dummy_pdf_response(f"matricula-{enrollment_id}.pdf")
class ScheduleExportView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        body = request.data or {}
        period = (body.get("academic_period") or "2025-I").strip()
        return ok(success=True,
                  downloadUrl=f"/api/schedules/export/pdf?academic_period={period}",
                  download_url=f"/api/schedules/export/pdf?academic_period={period}")
class ScheduleExportPDFView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        period = (request.query_params.get("academic_period") or "2025-I").strip()
        return _dummy_pdf_response(f"horario-{period}.pdf")
# ─────────────────────── Teachers / Classrooms ───────────────────────
class TeachersViewSet(viewsets.ViewSet):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def list(self, request, *args, **kwargs):
        qs = list_teacher_users_qs()
        if not qs.exists():
            ts = Teacher.objects.select_related("user").all()
            teachers = [{
                "id": t.user_id,
                "full_name": _get_full_name(t.user) if t.user else f"Teacher #{t.id}",
                "email": getattr(t.user, "email", "") if t.user else "",
                "username": getattr(t.user, "username", "") if t.user else "",
            } for t in ts]
            return ok(teachers=teachers)
        teachers = [{
            "id": u.id,
            "full_name": _get_full_name(u),
            "email": getattr(u, "email", "") or "",
            "username": getattr(u, "username", "") or "",
        } for u in qs]
        return ok(teachers=teachers)
class ClassroomsViewSet(viewsets.ReadOnlyModelViewSet):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    queryset = Classroom.objects.all()
    serializer_class = ClassroomSerializer
    def list(self, request, *args, **kwargs):
        data = self.get_serializer(self.get_queryset(), many=True).data
        return ok(classrooms=data)
# ───────────────────────── Plans / Mallas ─────────────────────────
class PlansViewSet(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    queryset = Plan.objects.select_related("career").all()
    def get_serializer_class(self):
        if self.action in ("create", "update", "partial_update"):
            return PlanCreateSerializer
        return PlanSerializer
    def list(self, request, *args, **kwargs):
        return ok(plans=PlanSerializer(self.get_queryset(), many=True).data)
    def create(self, request, *args, **kwargs):
        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        plan = ser.save()
        return ok(plan=PlanSerializer(plan).data)
    def retrieve(self, request, pk=None):
        plan = self.get_object()
        return ok(plan=PlanSerializer(plan).data)
    def update(self, request, pk=None):
        plan = self.get_object()
        ser = self.get_serializer(plan, data=request.data, partial=False)
        ser.is_valid(raise_exception=True)
        plan = ser.save()
        return ok(plan=PlanSerializer(plan).data)
    def destroy(self, request, pk=None):
        plan = self.get_object()
        try:
            with transaction.atomic():
                SectionScheduleSlot.objects.filter(section__plan_course__plan=plan).delete()
                AttendanceRow.objects.filter(session__section__plan_course__plan=plan).delete()
                AttendanceSession.objects.filter(section__plan_course__plan=plan).delete()
                Syllabus.objects.filter(section__plan_course__plan=plan).delete()
                EvaluationConfig.objects.filter(section__plan_course__plan=plan).delete()
                SectionGrades.objects.filter(section__plan_course__plan=plan).delete()
                Section.objects.filter(plan_course__plan=plan).delete()
                CoursePrereq.objects.filter(plan_course__plan=plan).delete()
                PlanCourse.objects.filter(plan=plan).delete()
                plan.delete()
            return ok(success=True)
        except IntegrityError:
            return Response(
                {"detail": "No se pudo eliminar el plan por restricciones de integridad (FK/PROTECT)."},
                status=status.HTTP_409_CONFLICT,
            )
        except Exception as e:
            return Response(
                {"detail": f"Error al eliminar plan: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
    @action(detail=True, methods=["get", "post"], url_path="courses")
    def courses(self, request, pk=None):
        plan = self.get_object()
        if request.method.lower() == "get":
            semester = request.query_params.get("semester")
            q = (request.query_params.get("q") or "").strip()
            base = PlanCourse.objects.filter(plan=plan).select_related("course")
            if q:
                base = base.filter(Q(course__code__icontains=q) | Q(course__name__icontains=q))
            if not semester:
                sems = (
                    base.values("semester")
                    .order_by("semester")
                    .annotate(total=Count("id"))
                )
                semesters = []
                for row in sems:
                    try:
                        sem = int(row.get("semester") or 0)
                    except Exception:
                        sem = 0
                    if sem <= 0:
                        continue
                    semesters.append({
                        "semester": sem,
                        "total": int(row.get("total") or 0),
                    })
                return ok(semesters=semesters)
            try:
                sem = int(semester)
            except Exception:
                return Response({"detail": "semester inválido"}, status=400)
            qs = base.filter(semester=sem).order_by("course__code", "id")
            return ok(
                semester=sem,
                total=qs.count(),
                courses=PlanCourseOutSerializer(qs, many=True).data
            )
        payload_ser = PlanCourseCreateSerializer(data=request.data)
        payload_ser.is_valid(raise_exception=True)
        data = payload_ser.validated_data
        with transaction.atomic():
            course, _ = Course.objects.get_or_create(
                code=data["code"],
                defaults={"name": data["name"], "credits": data.get("credits", 3)},
            )
            course.name = data["name"]
            course.credits = data.get("credits", course.credits)
            course.save()
            pc, _ = PlanCourse.objects.get_or_create(plan=plan, course=course)
            pc.semester = data.get("semester", pc.semester)
            pc.weekly_hours = data.get("weekly_hours", pc.weekly_hours)
            pc.type = data.get("type", pc.type)
            pc.save()
        return ok(course=PlanCourseOutSerializer(pc).data)
    @action(detail=True, methods=["put", "delete"], url_path=r"courses/(?P<course_id>\d+)")
    def course_detail(self, request, pk=None, course_id=None):
        plan = self.get_object()
        pc = get_object_or_404(PlanCourse, plan=plan, id=int(course_id))
        if request.method.lower() == "delete":
            pc.delete()
            return ok(success=True)
        body = request.data or {}
        with transaction.atomic():
            if "semester" in body:
                pc.semester = int(body["semester"])
            if "weekly_hours" in body:
                pc.weekly_hours = int(body["weekly_hours"])
            if "type" in body:
                pc.type = body["type"]
            pc.save()
            c = pc.course
            if "code" in body and body["code"] != c.code:
                c.code = body["code"]
            if "name" in body:
                c.name = body["name"]
            if "credits" in body:
                c.credits = int(body["credits"])
            c.save()
        return ok(course=PlanCourseOutSerializer(pc).data)
    @action(detail=True, methods=["get", "put"], url_path=r"courses/(?P<course_id>\d+)/prereqs")
    def prereqs(self, request, pk=None, course_id=None):
        plan = self.get_object()
        pc = get_object_or_404(PlanCourse, plan=plan, id=int(course_id))
        if request.method.lower() == "get":
            ids = CoursePrereq.objects.filter(plan_course=pc).values_list("prerequisite_id", flat=True)
            return ok(prerequisites=[{"id": i} for i in ids])
        ids = (request.data or {}).get("prerequisites", [])
        if not isinstance(ids, list):
            return Response({"detail": "prerequisites debe ser lista"}, status=400)
        valid = set(PlanCourse.objects.filter(plan=plan, id__in=ids).values_list("id", flat=True))
        with transaction.atomic():
            CoursePrereq.objects.filter(plan_course=pc).delete()
            for pid in valid:
                if pid == pc.id:
                    continue
                CoursePrereq.objects.create(plan_course=pc, prerequisite_id=pid)
        return ok(success=True, prerequisites=[{"id": i} for i in sorted(valid) if i != pc.id])
# ─────────────────────── Secciones / Horarios ───────────────────────
class SectionsViewSet(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    queryset = Section.objects.all()
    def get_queryset(self):
        qs = Section.objects.select_related(
            "plan_course__course", "teacher__user", "classroom"
        ).prefetch_related("schedule_slots")
        period = self.request.query_params.get("period")
        if period:
            qs = qs.filter(period=period)
        return qs
    def list(self, request, *args, **kwargs):
        return ok(sections=SectionOutSerializer(self.get_queryset(), many=True).data)
    def retrieve(self, request, pk=None, *args, **kwargs):
        sec = get_object_or_404(self.get_queryset(), id=int(pk))
        return ok(section=SectionOutSerializer(sec).data)
    def create(self, request, *args, **kwargs):
        ser = SectionCreateUpdateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        pc = get_object_or_404(PlanCourse, id=int(data["course_id"]))
        teacher = resolve_teacher(data.get("teacher_id"))
        room = resolve_classroom(data.get("room_id"))
        capacity = int(data.get("capacity") or 30)
        if room and room.capacity and capacity > room.capacity:
            return Response({"detail": "capacity excede la capacidad del aula"}, status=400)
        sec = Section.objects.create(
            plan_course=pc,
            teacher=teacher,
            classroom=room,
            capacity=capacity,
            period=(data.get("period") or "2025-I"),
            label=(data.get("label") or "A"),
        )
        slots = data.get("slots", [])
        for sl in slots:
            SectionScheduleSlot.objects.create(
                section=sec,
                weekday=DAY_TO_INT[sl["day"]],
                start=sl["start"],
                end=sl["end"],
            )
        sec = self.get_queryset().get(id=sec.id)
        return ok(section=SectionOutSerializer(sec).data)
    def update(self, request, pk=None, *args, **kwargs):
        sec = get_object_or_404(Section, id=int(pk))
        ser = SectionCreateUpdateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        if "teacher_id" in data:
            sec.teacher = resolve_teacher(data.get("teacher_id"))
        if "room_id" in data:
            sec.classroom = resolve_classroom(data.get("room_id"))
        if "capacity" in data:
            sec.capacity = int(data.get("capacity") or sec.capacity)
        if "period" in data:
            sec.period = data.get("period") or sec.period
        if "label" in data:
            sec.label = data.get("label") or sec.label
        if sec.classroom and sec.classroom.capacity and sec.capacity > sec.classroom.capacity:
            return Response({"detail": "capacity excede la capacidad del aula"}, status=400)
        sec.save()
        if "slots" in data:
            SectionScheduleSlot.objects.filter(section=sec).delete()
            for sl in data.get("slots", []):
                SectionScheduleSlot.objects.create(
                    section=sec,
                    weekday=DAY_TO_INT[sl["day"]],
                    start=sl["start"],
                    end=sl["end"],
                )
        sec = self.get_queryset().get(id=sec.id)
        return ok(section=SectionOutSerializer(sec).data)
    def partial_update(self, request, pk=None, *args, **kwargs):
        sec = get_object_or_404(Section, id=int(pk))
        body = request.data or {}
        if "teacher_id" in body:
            sec.teacher = resolve_teacher(body.get("teacher_id"))
        if "room_id" in body:
            sec.classroom = resolve_classroom(body.get("room_id"))
        if "capacity" in body:
            sec.capacity = int(body.get("capacity") or sec.capacity)
        if "period" in body:
            sec.period = body.get("period") or sec.period
        if "label" in body:
            sec.label = body.get("label") or sec.label
        if sec.classroom and sec.classroom.capacity and sec.capacity > sec.classroom.capacity:
            return Response({"detail": "capacity excede la capacidad del aula"}, status=400)
        sec.save()
        if "slots" in body and isinstance(body["slots"], list):
            tmp = SectionCreateUpdateSerializer(data={"course_id": sec.plan_course_id, "slots": body["slots"]})
            tmp.is_valid(raise_exception=True)
            SectionScheduleSlot.objects.filter(section=sec).delete()
            for sl in tmp.validated_data.get("slots", []):
                SectionScheduleSlot.objects.create(
                    section=sec,
                    weekday=DAY_TO_INT[sl["day"]],
                    start=sl["start"],
                    end=sl["end"],
                )
        sec = self.get_queryset().get(id=sec.id)
        return ok(section=SectionOutSerializer(sec).data)
    def destroy(self, request, pk=None, *args, **kwargs):
        sec = get_object_or_404(Section, id=int(pk))
        sec.delete()
        return ok(success=True)
    @action(detail=True, methods=["get", "put"], url_path="schedule")
    def schedule(self, request, pk=None):
        sec = get_object_or_404(Section, id=int(pk))
        if request.method.lower() == "get":
            sec = Section.objects.prefetch_related("schedule_slots").get(id=sec.id)
            return ok(slots=SectionOutSerializer(sec).data["slots"])
        slots = (request.data or {}).get("slots", [])
        if not isinstance(slots, list):
            return Response({"detail": "slots debe ser lista"}, status=400)
        tmp = SectionCreateUpdateSerializer(data={"course_id": sec.plan_course_id, "slots": slots})
        tmp.is_valid(raise_exception=True)
        SectionScheduleSlot.objects.filter(section=sec).delete()
        for sl in tmp.validated_data.get("slots", []):
            SectionScheduleSlot.objects.create(
                section=sec,
                weekday=DAY_TO_INT[sl["day"]],
                start=sl["start"],
                end=sl["end"],
            )
        sec = Section.objects.prefetch_related("schedule_slots").get(id=sec.id)
        return ok(success=True, slots=SectionOutSerializer(sec).data["slots"])
class SectionsScheduleConflictsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        payload = request.data or {}
        slots = payload.get("slots", [])
        conflicts = []
        seen = set()
        for sl in slots:
            key = (sl.get("day") or sl.get("weekday"), sl.get("start"), sl.get("end"))
            if key in seen:
                conflicts.append({"message": f"Conflicto en {key[0]} {key[1]}-{key[2]}"})
            seen.add(key)
        return ok(conflicts=conflicts)
# ─────────────────────── Attendance ───────────────────────
class AttendanceSessionsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, section_id):
        qs = AttendanceSession.objects.filter(section_id=section_id).prefetch_related("rows").order_by("-id")
        return ok(sessions=AttendanceSessionSerializer(qs, many=True).data)
    def post(self, request, section_id):
        body = request.data or {}
        date_str = body.get("date")
        if date_str:
            try:
                d = datetime.strptime(date_str, "%Y-%m-%d").date()
            except Exception:
                return Response({"detail": "date inválida (YYYY-MM-DD)"}, status=400)
            sess, _ = AttendanceSession.objects.get_or_create(section_id=section_id, date=d)
        else:
            sess = AttendanceSession.objects.create(section_id=section_id)
        return ok(session=AttendanceSessionSerializer(sess).data)
class AttendanceSessionCloseView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request, section_id, session_id):
        sess = get_object_or_404(AttendanceSession, id=session_id, section_id=section_id)
        sess.closed = True
        sess.save()
        return ok(success=True)
class AttendanceSessionSetView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def put(self, request, section_id, session_id):
        sess = get_object_or_404(AttendanceSession, id=session_id, section_id=section_id)
        rows = (request.data or {}).get("rows", [])
        if not isinstance(rows, list):
            return Response({"detail": "rows debe ser lista"}, status=400)
        with transaction.atomic():
            AttendanceRow.objects.filter(session=sess).delete()
            for r in rows:
                sid = r.get("student_id")
                try:
                    sid = int(sid)
                except Exception:
                    sid = 0
                st = (r.get("status") or "").upper().strip()
                AttendanceRow.objects.create(session=sess, student_id=sid, status=st)
        return ok(success=True)
# ─────────────────────── Syllabus ───────────────────────
class SyllabusView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, section_id):
        s = Syllabus.objects.filter(section_id=section_id).first()
        if not s:
            return ok(syllabus=None)
        return ok(syllabus={"filename": s.file.name, "size": getattr(s.file, "size", 0)})
    def post(self, request, section_id):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "Archivo requerido"}, status=status.HTTP_400_BAD_REQUEST)
        obj, _ = Syllabus.objects.get_or_create(section_id=section_id)
        obj.file = f
        obj.save()
        return ok(syllabus={"filename": obj.file.name, "size": getattr(obj.file, "size", 0)})
    def delete(self, request, section_id):
        Syllabus.objects.filter(section_id=section_id).delete()
        return ok(success=True)
# ─────────────────────── Evaluación config ───────────────────────
class EvaluationConfigView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, section_id):
        obj = EvaluationConfig.objects.filter(section_id=section_id).first()
        return ok(config=(obj.config if obj else []))
    def put(self, request, section_id):
        cfg = request.data if isinstance(request.data, list) else (request.data or {}).get("config", [])
        obj, _ = EvaluationConfig.objects.get_or_create(section_id=section_id)
        obj.config = cfg
        obj.save()
        return ok(config=obj.config)
# ─────────────────────────── Kardex ───────────────────────────
class KardexView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def _final_grade(self, row: dict):
        if not isinstance(row, dict):
            return None
        for k in ("FINAL", "final", "PROMEDIO", "promedio", "AVERAGE", "average"):
            v = row.get(k)
            try:
                if v is not None and str(v).strip() != "":
                    return float(v)
            except Exception:
                pass
        vals = []
        for k in ("PARCIAL_1", "PARCIAL_2", "PARCIAL_3", "P1", "P2", "P3"):
            v = row.get(k)
            try:
                if v is not None and str(v).strip() != "":
                    vals.append(float(v))
            except Exception:
                pass
        if vals:
            return sum(vals) / len(vals)
        return None
    def _status_from_grade(self, g):
        if g is None:
            return "SIN NOTA"
        return "APROBADO" if g >= 11 else "DESAPROBADO"
    def get(self, request, student_id):
        st = None
        doc = str(student_id).strip()
        if doc.isdigit():
            st = StudentProfile.objects.filter(num_documento=doc).first()
            if not st:
                st = StudentProfile.objects.filter(id=int(doc)).first()
        else:
            st = StudentProfile.objects.filter(num_documento=doc).first()
        if not st:
            return ok(student_id=student_id, student_name="No encontrado", career_name="", credits_earned=0, gpa=None, items=[])
        student_name = f"{st.apellido_paterno} {st.apellido_materno} {st.nombres}".strip()
        career_name = st.programa_carrera or ""
        best_by_key = {}
        recs = (
            AcademicGradeRecord.objects
            .select_related("course")
            .filter(student=st)
            .order_by("-id")
        )
        for rec in recs:
            crs = rec.course
            period = (rec.term or "").strip()
            course_code = (crs.code or "").strip() or f"CRS-{crs.id}"
            semester = 0
            weekly_hours = 0
            if st.plan_id:
                pc = PlanCourse.objects.filter(plan_id=st.plan_id, course=crs).first()
                if pc:
                    semester = int(pc.semester or 0)
                    weekly_hours = int(pc.weekly_hours or 0)
            grade = float(rec.final_grade) if rec.final_grade is not None else None
            status_txt = self._status_from_grade(grade)
            credits = int(getattr(crs, "credits", 0) or 0)
            item = {
                "period": period,
                "semester": semester,
                "weekly_hours": weekly_hours,
                "course_code": course_code,
                "course_name": crs.name,
                "credits": credits,
                "grade": grade,
                "status": status_txt,
            }
            key = (period, course_code)
            def score(x):
                g = x.get("grade")
                return -1 if g is None else float(g)
            prev = best_by_key.get(key)
            if (prev is None) or (score(item) > score(prev)):
                best_by_key[key] = item
        items = list(best_by_key.values())
        items.sort(key=lambda x: (str(x.get("period") or ""), int(x.get("semester") or 0), str(x.get("course_code") or "")))
        credits_earned = 0
        sum_w = 0.0
        sum_c = 0
        for it in items:
            cr = int(it.get("credits") or 0)
            g = it.get("grade")
            if it.get("status") == "APROBADO":
                credits_earned += cr
            if g is not None and cr > 0:
                sum_w += float(g) * cr
                sum_c += cr
        gpa = round(sum_w / sum_c, 2) if sum_c > 0 else None
        return ok(
            student_id=st.num_documento,
            student_name=student_name,
            career_name=career_name,
            credits_earned=credits_earned,
            gpa=gpa,
            items=items,
        )
class KardexExportXlsxView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, student_id):
        period_q = (request.query_params.get("period") or "").strip()
        kv = KardexView()
        kv.request = request
        resp = kv.get(request, student_id)
        data = getattr(resp, "data", None) or {}
        items = data.get("items") or []
        if period_q:
            items = [x for x in items if str(x.get("period") or "").strip() == period_q]
        template_path = os.path.join(settings.BASE_DIR, "academic", "templates", "kardex_template.xlsx")
        wb = load_workbook(template_path)
        ws = wb[wb.sheetnames[0]]
        start_row = 14
        for r in range(start_row, start_row + 200):
            for c in range(1, 10):
                ws.cell(r, c).value = None
        today = timezone.now().date().isoformat()
        def s_int(v):
            try:
                return int(v)
            except Exception:
                return 0
        def s_float(v):
            try:
                return float(v)
            except Exception:
                return None
        items_sorted = sorted(items, key=lambda x: (s_int(x.get("semester") or 0), str(x.get("course_code") or "")))
        for idx, it in enumerate(items_sorted, start=1):
            r = start_row + (idx - 1)
            semester = s_int(it.get("semester") or 0) or ""
            hours = s_int(it.get("weekly_hours") or 0) or ""
            credits = s_int(it.get("credits") or 0) or ""
            grade = s_float(it.get("grade"))
            status_txt = it.get("status") or ""
            ws.cell(r, 1).value = semester
            ws.cell(r, 2).value = idx
            ws.cell(r, 3).value = it.get("course_name")
            ws.cell(r, 4).value = hours
            ws.cell(r, 5).value = credits
            ws.cell(r, 6).value = "" if grade is None else grade
            ws.cell(r, 7).value = today
            ws.cell(r, 8).value = status_txt
            ws.cell(r, 9).value = "" if grade is None else grade
        filename = f"kardex-{student_id}{('-' + period_q) if period_q else ''}.xlsx"
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return FileResponse(
            bio,
            as_attachment=True,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
class KardexBoletaPDFView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, student_id):
        st = _student_lookup(student_id)
        if not st:
            return Response({"detail": "Estudiante no encontrado"}, status=404)
        student_name = f"{st.apellido_paterno} {st.apellido_materno} {st.nombres}".strip()
        student_code = st.num_documento
        career_name = (st.programa_carrera or "").upper()
        grouped = build_boleta_full(st)
        page1 = [x for x in grouped if 1 <= int(x["semester"]) <= 5]
        page2 = [x for x in grouped if 6 <= int(x["semester"]) <= 10]
        logo_data, sig_data = _get_institution_media_datauris(request)
        ctx = {
            "student_name": student_name,
            "student_code": student_code,
            "career_name": career_name,

            "logo_url": logo_data,
            "signature_url": sig_data,

            "page1": page1,
            "page2": page2,
            "period_label": "",
        }
        html = render_to_string("kardex/boleta_comunicacion.html", ctx)
        pdf_bytes = html_to_pdf_bytes(html)
        filename = f"boleta-{student_id}.pdf"
        return HttpResponse(pdf_bytes, content_type="application/pdf",
                            headers={"Content-Disposition": f'attachment; filename="{filename}"'})
class KardexConstanciaPDFView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, student_id):
        kv = KardexView()
        kv.request = request
        resp = kv.get(request, student_id)
        data = getattr(resp, "data", None) or {}
        student_name = data.get("student_name", "") or ""
        student_code = data.get("student_id", "") or ""
        career_name = data.get("career_name", "") or ""
        template_name = _pick_kardex_template(career_name)
        template_path = os.path.join(settings.BASE_DIR, "templates", "kardex", template_name)
        if not os.path.exists(template_path):
            template_path = os.path.join(settings.BASE_DIR, "templates", "kardex", "inicial.pdf")
        tpl_pages = len(PdfReader(template_path).pages)
        def draw_fn(c, page_i):
            if page_i == 0:
                _draw_text(c, *KARDEX_POS["name"], student_name, size=10, bold=True)
                _draw_text(c, *KARDEX_POS["code"], student_code, size=9, bold=False)
        overlay_reader = _make_overlay_pdf(tpl_pages, draw_fn)
        pdf_bytes = _merge_overlay(template_path, overlay_reader)
        filename = f"constancia-{student_id}.pdf"
        return HttpResponse(
            pdf_bytes,
            content_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    def post(self, request, student_id):
        return self.get(request, student_id)
class KardexBoletaPeriodoPDFView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, student_id):
        period_q = (request.query_params.get("period") or "").strip()
        if not period_q:
            return Response({"detail": "Falta period (ej: 2018-II)"}, status=400)

        st = _student_lookup(student_id)
        if not st:
            return Response({"detail": "Estudiante no encontrado"}, status=404)

        pq = _norm_term(period_q)

        try:
            pdf_bytes = _render_reporte_periodo_pdf(request, st, pq)
            filename = f"reporte-calificaciones-{student_id}-{pq}.pdf"
            return HttpResponse(
                pdf_bytes,
                content_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        except ValueError as ve:
            # viene de "no hay registros"
            logger.info("Sin registros PDF periodo student=%s period=%s", student_id, pq)
            return Response({"detail": "No hay registros", "error": str(ve)}, status=404)
        except Exception as e:
            logger.exception("Error PDF periodo student=%s period=%s", student_id, pq)
            return Response({"detail": "Error interno generando PDF", "error": str(e)}, status=500)
class KardexBoletaAnioPDFView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, student_id):
        period_q = (request.query_params.get("period") or "").strip()
        if not period_q:
            return Response({"detail": "Falta period (ej: 2018-I)"}, status=400)

        st = _student_lookup(student_id)
        if not st:
            return Response({"detail": "Estudiante no encontrado"}, status=404)

        p = _norm_term(period_q)
        m = re.match(r"^(\d{4})-(I|II|1|2)$", p)
        if not m:
            return Response({"detail": "period inválido (ej: 2018-I / 2018-II)"}, status=400)

        year = int(m.group(1))
        periods = [f"{year}-I", f"{year}-II"]

        writer = PdfWriter()
        any_pages = False

        try:
            for per in periods:
                pq = _norm_term(per)
                ctx, err = _build_reporte_periodo_ctx(request, st, pq)
                if err:
                    continue

                html = render_to_string("kardex/reporte_calificaciones.html", ctx)
                pdf_bytes = html_to_pdf_bytes(html)

                reader = PdfReader(BytesIO(pdf_bytes))
                for pg in reader.pages:
                    writer.add_page(pg)
                    any_pages = True

            if not any_pages:
                return Response({"detail": f"No hay reportes para el año {year}."}, status=404)

            out = BytesIO()
            writer.write(out)
            out.seek(0)

            filename = f"reporte-calificaciones-{student_id}-{year}.pdf"
            return HttpResponse(
                out.getvalue(),
                content_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        except Exception as e:
            return Response({"detail": "Error interno generando PDF", "error": str(e)}, status=500)
    
# ───────────────────── Procesos académicos ─────────────────────
class ProcessesCreateView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request, ptype=None):
        body = request.data or {}
        proc = AcademicProcess.objects.create(
            kind=ptype,
            student_id=body.get("student_id") or 0,
            status="PENDIENTE",
            note=body.get("reason", "") or "",
        )
        return ok(process={
            "id": proc.id,
            "type": proc.kind,
            "status": proc.status,
            "student_id": proc.student_id,
            "period": body.get("period"),
            "reason": body.get("reason", ""),
            "extra": body.get("extra", ""),
            "created_at": datetime.now().isoformat(),
        })
class ProcessesListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        qs = AcademicProcess.objects.all().order_by("-id")
        data = [{"id": p.id, "type": p.kind, "status": p.status, "student_id": p.student_id, "note": p.note} for p in qs]
        return ok(processes=data)
class ProcessesMineView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        qs = AcademicProcess.objects.all().order_by("-id")
        data = [{"id": p.id, "type": p.kind, "status": p.status, "student_id": p.student_id, "note": p.note} for p in qs]
        return ok(processes=data)
class ProcessDetailView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, pid):
        p = get_object_or_404(AcademicProcess, id=pid)
        return ok(process={"id": p.id, "type": p.kind, "status": p.status, "student_id": p.student_id, "note": p.note})
class ProcessStatusView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request, pid):
        body = request.data or {}
        p = get_object_or_404(AcademicProcess, id=pid)
        if body.get("status"):
            p.status = body["status"]
        if body.get("note") is not None:
            p.note = body.get("note") or ""
        p.save()
        return ok(success=True)
class ProcessNotifyView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request, pid):
        return ok(sent=True)
class ProcessFilesView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, pid):
        qs = ProcessFile.objects.filter(process_id=pid).order_by("-id")
        files = [{"id": f.id, "name": f.file.name, "size": getattr(f.file, "size", 0)} for f in qs]
        return ok(files=files)
    def post(self, request, pid):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "Archivo requerido"}, status=400)
        pf = ProcessFile.objects.create(process_id=pid, file=f, note=(request.data or {}).get("note", ""))
        return ok(file={"id": pf.id, "name": pf.file.name, "size": getattr(pf.file, "size", 0)})
class ProcessFileDeleteView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def delete(self, request, pid, file_id):
        ProcessFile.objects.filter(process_id=pid, id=file_id).delete()
        return ok(success=True)
# ───────────────────── Reportes académicos ─────────────────────
class AcademicReportsSummaryView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        students_count = list_student_users_qs().count()
        return ok(summary={
            "students": students_count,
            "sections": Section.objects.count(),
            "teachers": count_teachers(),
            "occupancy": 0.76,
            "avg_gpa": 13.4
        })
def _xlsx_response(filename: str, rows, sheet_name="Reporte"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    for r in rows:
        ws.append(r)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return FileResponse(
        bio,
        as_attachment=True,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
class AcademicCareersListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        qs = (
            Plan.objects
            .select_related("career")
            .exclude(career__isnull=True)
            .values("career_id", "career__name")
            .distinct()
            .order_by("career__name")
        )
        careers = [{"id": r["career_id"], "name": r["career__name"]} for r in qs if r["career_id"]]
        return ok(careers=careers)
class AcademicReportPerformanceXlsxView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        period = (request.query_params.get("period") or "").strip()
        career_id = (request.query_params.get("career_id") or "").strip()
        date_from = (request.query_params.get("from") or "").strip()
        date_to = (request.query_params.get("to") or "").strip()
        rows = [["Métrica", "Valor"]]
        rows += [
            ["Periodo", period or "ALL"],
            ["Carrera", career_id or "ALL"],
            ["Desde", date_from or ""],
            ["Hasta", date_to or ""],
            ["Total secciones", Section.objects.count()],
            ["Total estudiantes (roles)", list_student_users_qs().count()],
            ["Generado", timezone.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]
        return _xlsx_response("performance.xlsx", rows, sheet_name="Performance")
class AcademicReportOccupancyXlsxView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        period = (request.query_params.get("period") or "").strip()
        career_id = (request.query_params.get("career_id") or "").strip()
        date_from = (request.query_params.get("from") or "").strip()
        date_to = (request.query_params.get("to") or "").strip()
        total_rooms = Classroom.objects.count()
        total_sections = Section.objects.count()
        rows = [["Métrica", "Valor"]]
        rows += [
            ["Periodo", period or "ALL"],
            ["Carrera", career_id or "ALL"],
            ["Desde", date_from or ""],
            ["Hasta", date_to or ""],
            ["Total aulas", total_rooms],
            ["Total secciones", total_sections],
            ["Ocupación estimada", 0.76],
            ["Generado", timezone.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]
        return _xlsx_response("occupancy.xlsx", rows, sheet_name="Occupancy")
# ───────────────────── Docente / Estudiantes / Notas / Acta / Import ─────────────────────
class TeacherSectionsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, teacher_user_id: int):
        teacher = resolve_teacher(teacher_user_id)
        if not teacher:
            return ok(sections=[])
        qs = (
            Section.objects
            .select_related("plan_course__course", "teacher__user", "classroom")
            .prefetch_related("schedule_slots")
            .filter(teacher=teacher)
            .order_by("-id")
        )
        sections = []
        for s in qs:
            crs = s.plan_course.course
            sections.append({
                "id": s.id,
                "course_name": crs.name,
                "course_code": crs.code,
                "section_code": s.label,
                "label": s.label,
                "period": s.period,
                "plan_course_id": s.plan_course_id,
                "room_name": s.classroom.code if s.classroom else "",
            })
        return ok(sections=sections)
class SectionStudentsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, section_id: int):
        qs = list_student_users_qs().order_by("id")[:200]
        students = []
        for u in qs:
            full = _get_full_name(u)
            parts = full.split()
            first = parts[0] if parts else ""
            last = " ".join(parts[1:]) if len(parts) > 1 else ""
            students.append({
                "id": u.id,
                "first_name": first,
                "last_name": last,
            })
        return ok(students=students)
class SectionGradesView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, section_id: int):
        sec = get_object_or_404(Section, id=section_id)
        bundle, _ = SectionGrades.objects.get_or_create(section=sec)
        return ok(grades=bundle.grades, submitted=bundle.submitted)
class GradesSaveView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        body = request.data or {}
        section_id = body.get("section_id")
        grades = body.get("grades") or {}
        if not section_id:
            return Response({"detail": "section_id requerido"}, status=400)
        if not isinstance(grades, dict):
            return Response({"detail": "grades debe ser objeto"}, status=400)
        sec = get_object_or_404(Section, id=int(section_id))
        bundle, _ = SectionGrades.objects.get_or_create(section=sec)
        if bundle.submitted:
            return Response({"detail": "Las calificaciones ya están cerradas."}, status=409)
        bundle.grades = grades
        bundle.save()
        return ok(success=True)
class GradesSubmitView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        body = request.data or {}
        section_id = body.get("section_id")
        grades = body.get("grades") or {}
        if not section_id:
            return Response({"detail": "section_id requerido"}, status=400)
        if not isinstance(grades, dict):
            return Response({"detail": "grades debe ser objeto"}, status=400)
        sec = get_object_or_404(Section, id=int(section_id))
        bundle, _ = SectionGrades.objects.get_or_create(section=sec)
        bundle.grades = grades
        bundle.submitted = True
        bundle.submitted_at = timezone.now()
        bundle.save()
        return ok(success=True, submitted=True)
class GradesReopenView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        if not user_has_any_role(request.user, ["REGISTRAR", "ADMIN_ACADEMIC"]):
            return Response({"detail": "No autorizado"}, status=403)
        body = request.data or {}
        section_id = body.get("section_id")
        if not section_id:
            return Response({"detail": "section_id requerido"}, status=400)
        sec = get_object_or_404(Section, id=int(section_id))
        bundle, _ = SectionGrades.objects.get_or_create(section=sec)
        bundle.submitted = False
        bundle.submitted_at = None
        bundle.save()
        return ok(success=True, submitted=False)
class SectionActaView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request, section_id: int):
        return ok(
            success=True,
            downloadUrl=f"/api/sections/{section_id}/acta/pdf",
            download_url=f"/api/sections/{section_id}/acta/pdf",
        )
class SectionActaPDFView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, section_id: int):
        return _dummy_pdf_response(f"acta-section-{section_id}.pdf")
class SectionActaQRView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request, section_id: int):
        return ok(
            success=True,
            qrUrl=f"/api/sections/{section_id}/acta/qr/png",
            qr_url=f"/api/sections/{section_id}/acta/qr/png",
        )
class SectionActaQRPngView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, section_id: int):
        png_1x1 = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMB/6X1fQAAAABJRU5ErkJggg=="
        )
        return HttpResponse(png_1x1, content_type="image/png")
# ─────────────────────── Attendance Import ───────────────────────
ALLOWED_ATT = {"PRESENT", "ABSENT", "LATE", "EXCUSED"}
class AttendanceImportPreviewView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        f = request.FILES.get("file")
        section_id = (request.data or {}).get("section_id")
        if not f or not section_id:
            return Response({"detail": "file y section_id son requeridos"}, status=400)
        try:
            int(section_id)
        except Exception:
            return Response({"detail": "section_id inválido"}, status=400)
        errors = []
        preview = []
        decoded = f.read().decode("utf-8-sig", errors="ignore").splitlines()
        reader = csv.DictReader(decoded)
        for idx, row in enumerate(reader, start=2):
            status_val = (row.get("status") or row.get("estado") or "").strip().upper()
            date_val = (row.get("date") or row.get("fecha") or "").strip()
            student_name = (row.get("student_name") or row.get("nombre") or row.get("student") or "").strip()
            student_id = (row.get("student_id") or row.get("id") or "").strip()
            if status_val not in ALLOWED_ATT:
                errors.append({"row": idx, "message": f"status inválido: {status_val}"})
                continue
            if date_val:
                try:
                    datetime.strptime(date_val, "%Y-%m-%d")
                except Exception:
                    errors.append({"row": idx, "message": "date inválida (YYYY-MM-DD)"})
                    continue
            else:
                date_val = str(timezone.now().date())
            if not student_id and not student_name:
                errors.append({"row": idx, "message": "Falta student_id o student_name"})
                continue
            preview.append({
                "student_id": int(student_id) if student_id.isdigit() else None,
                "student_name": student_name or (f"ID {student_id}" if student_id else "Desconocido"),
                "date": date_val,
                "status": status_val,
            })
        return ok(preview=preview, errors=errors)
class AttendanceImportSaveView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        body = request.data or {}
        section_id = body.get("section_id")
        data = body.get("attendance_data") or []
        if not section_id:
            return Response({"detail": "section_id requerido"}, status=400)
        if not isinstance(data, list):
            return Response({"detail": "attendance_data debe ser lista"}, status=400)
        section = get_object_or_404(Section, id=int(section_id))
        by_date = {}
        for r in data:
            dt = (r.get("date") or "").strip()
            st = (r.get("status") or "").strip().upper()
            sid = r.get("student_id")
            if not dt or st not in ALLOWED_ATT:
                continue
            by_date.setdefault(dt, []).append((sid, st))
        with transaction.atomic():
            for dt, rows in by_date.items():
                d = datetime.strptime(dt, "%Y-%m-%d").date()
                sess, _ = AttendanceSession.objects.get_or_create(section=section, date=d)
                AttendanceRow.objects.filter(session=sess).delete()
                for sid, st in rows:
                    try:
                        sid_int = int(sid) if sid is not None else 0
                    except Exception:
                        sid_int = 0
                    AttendanceRow.objects.create(session=sess, student_id=sid_int, status=st)
        return ok(success=True)
# =========================
# BOLETA OFICIAL (HTML->PDF)
# =========================
SEM_LABELS = {
    1: "PRIMERO", 2: "SEGUNDO", 3: "TERCERO", 4: "CUARTO", 5: "QUINTO",
    6: "SEXTO", 7: "SEPTIMO", 8: "OCTAVO", 9: "NOVENO", 10: "DECIMO",
}
def _norm_text(s: str) -> str:
    s = "" if s is None else str(s)
    s = s.strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"\s+", " ", s)
    return s.lower()
def _norm_term(s: str) -> str:
    s = "" if s is None else str(s)
    s = s.strip().upper()
    s = re.sub(r"\s+", "", s)
    s = s.replace("/", "-")
    return s
def _student_lookup(student_id: str):
    doc = str(student_id).strip()
    if doc.isdigit():
        return (
            StudentProfile.objects.filter(num_documento=doc).first()
            or StudentProfile.objects.filter(id=int(doc)).first()
        )
    return StudentProfile.objects.filter(num_documento=doc).first()
def _resolve_plan_for_student(student: StudentProfile):
    if getattr(student, "plan_id", None):
        return student.plan_id
    career_name = (getattr(student, "programa_carrera", "") or "").strip()
    if not career_name:
        return None
    car = Career.objects.filter(name__iexact=career_name).first() or Career.objects.filter(name__icontains=career_name).first()
    if not car:
        return None
    plan = Plan.objects.filter(career=car).order_by("id").first()
    if not plan:
        return None
    try:
        student.plan_id = plan.id
        student.save(update_fields=["plan"])
    except Exception:
        pass
    return plan.id
def _plan_pc_map_by_name(plan_id: int):
    pc_by_name = {}
    qs = PlanCourse.objects.select_related("course").filter(plan_id=plan_id)
    for pc in qs:
        key = _norm_text(getattr(pc.course, "name", "") or "")
        if key and key not in pc_by_name:
            pc_by_name[key] = pc
    return pc_by_name
def _boleta_group_from_plan_courses(pcs):
    by_sem = {}
    for pc in pcs:
        sem = int(getattr(pc, "semester", 0) or 0)
        if sem <= 0:
            continue
        by_sem.setdefault(sem, []).append(pc)
    grouped = []
    for sem in sorted(by_sem.keys()):
        pcs_sem = sorted(
            by_sem[sem],
            key=lambda x: (getattr(x.course, "code", "") or "", getattr(x.course, "name", "") or "", x.id),
        )
        rows = []
        total_credits = 0
        for idx, pc in enumerate(pcs_sem, start=1):
            cr = int(getattr(pc.course, "credits", 0) or 0)
            hrs = int(getattr(pc, "weekly_hours", 0) or 0)
            total_credits += cr
            rows.append({
                "n": idx,
                "course_name": pc.course.name or "",
                "hours": hrs if hrs else "",
                "credits": cr if cr else "",
                "grade": "",
            })
        grouped.append({
            "semester": sem,
            "label": SEM_LABELS.get(sem, f"SEM {sem}"),
            "rows": rows,
            "rowspan": max(1, len(rows)),
            "total_credits": total_credits,
        })
    return grouped
def _apply_grades_to_grouped(grouped, grade_by_normname: dict):
    for sem in grouped:
        for row in sem["rows"]:
            key = _norm_text(row["course_name"])
            g = grade_by_normname.get(key)
            row["grade"] = "" if g is None else (int(g) if float(g).is_integer() else round(float(g), 2))
def _grades_map_for_student(student: StudentProfile, period_q: str = ""):
    pq = _norm_term(period_q) if period_q else ""
    recs = list(AcademicGradeRecord.objects.select_related("course").filter(student=student))
    if pq:
        recs = [r for r in recs if _norm_term(getattr(r, "term", "") or "") == pq]
    grade_by_name = {}
    for r in recs:
        name = _norm_text(getattr(r.course, "name", "") or "")
        try:
            g = None if r.final_grade is None else float(r.final_grade)
        except Exception:
            g = None
        if not name:
            continue
        prev = grade_by_name.get(name)
        if prev is None or (g is not None and (prev is None or g > prev)):
            grade_by_name[name] = g
    return grade_by_name
def build_boleta_full(student: StudentProfile):
    plan_id = _resolve_plan_for_student(student)
    if not plan_id:
        return []
    pcs = list(PlanCourse.objects.select_related("course").filter(plan_id=plan_id).order_by("semester", "id"))
    grouped = _boleta_group_from_plan_courses(pcs)
    _apply_grades_to_grouped(grouped, _grades_map_for_student(student, period_q=""))
    return grouped
def build_boleta_for_period(student: StudentProfile, period_q: str):
    pq = _norm_term(period_q)
    if not pq:
        return []
    plan_id = _resolve_plan_for_student(student)
    if not plan_id:
        return []
    recs = list(AcademicGradeRecord.objects.select_related("course").filter(student=student))
    recs = [r for r in recs if _norm_term(getattr(r, "term", "") or "") == pq]
    if not recs:
        return []
    course_ids = list({r.course_id for r in recs})
    pcs = list(PlanCourse.objects.select_related("course").filter(plan_id=plan_id, course_id__in=course_ids))
    # Fallback por nombre
    pc_by_name = _plan_pc_map_by_name(plan_id)
    found = set(_norm_text(pc.course.name) for pc in pcs if pc.course and pc.course.name)
    for r in recs:
        k = _norm_text(getattr(r.course, "name", "") or "")
        if k and k not in found and k in pc_by_name:
            pcs.append(pc_by_name[k])
            found.add(k)
    # ✅ Fallback extremo: si aún no hay pcs, usa recs directamente
    if not pcs:
        grouped = [{
            "semester": 0,
            "label": f"PERIODO {pq}",
            "rows": [],
            "rowspan": 1,
            "total_credits": "",
        }]
        for idx, r in enumerate(recs, start=1):
            name = getattr(r.course, "name", "") or ""
            try:
                g = None if r.final_grade is None else float(r.final_grade)
            except Exception:
                g = None
            grouped[0]["rows"].append({
                "n": idx,
                "course_name": name,
                "hours": "",
                "credits": "",
                "grade": "" if g is None else (int(g) if float(g).is_integer() else round(float(g), 2)),
            })
        grouped[0]["rowspan"] = max(1, len(grouped[0]["rows"]))
        return grouped
    grouped = _boleta_group_from_plan_courses(pcs)
    _apply_grades_to_grouped(grouped, _grades_map_for_student(student, period_q=period_q))
    return grouped

class CoursesListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        q = (request.query_params.get("q") or "").strip()
        qs = Course.objects.all().order_by("code", "name")
        if q:
            qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q))

        items = list(qs.values("id", "code", "name", "credits")[:1000])
        return Response({"items": items})